"""
Excel formula builders and the shared report layout.

Every reported figure in a Flux pack is a formula over the input sheet, so the
workbook recalculates when an input is edited. This module builds those formula
strings and defines the column layout every reporting sheet uses; it holds no
styling and touches no worksheet.

Three lever cells live on the P&L sheet - the two materiality floors and the
number of months elapsed. Every other sheet points at those cells, so changing
one number re-flags or re-projects the whole pack.

The layout is fixed for every reporting sheet:

    <label columns> | Month Act Bud Var Var% | YTD Act Bud Var Var%
                    | FY Budget | Run rate FY | Var to FY | F/U | Flag

Sheets differ only in how many label columns they carry and in what the measure
means, so a reader who has learnt one sheet has learnt all of them. The column
letters are derived from the label count rather than hand-counted, because a
sheet with three label columns and one with a single label column were the two
places the old layout drifted apart.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Lever cells
# ---------------------------------------------------------------------------
LEVER_EUR = "C9"
LEVER_PCT = "G9"
LEVER_MONTHS = "K9"
LEV_E = "$C$9"
LEV_P = "$G$9"
LEV_M = "$K$9"

# Absolute references for sheets other than the P&L, which owns the levers.
PNL_SHEET = "P&L Report"
PL_E = f"'{PNL_SHEET}'!{LEV_E}"
PL_P = f"'{PNL_SHEET}'!{LEV_P}"
PL_M = f"'{PNL_SHEET}'!{LEV_M}"

DEFAULT_ABS_THRESHOLD = 25_000
DEFAULT_PCT_THRESHOLD = 0.10


# ---------------------------------------------------------------------------
# "Not meaningful" thresholds - mirrored from engine.py so the workbook and the
# Python engine reach the same verdict on the same numbers.
# ---------------------------------------------------------------------------
NM_RATIO_CAP = 10          # variance above 1,000% of the base
NM_ABS_FLOOR = 100         # base below this counts as zero


# ---------------------------------------------------------------------------
# The shared column layout
# ---------------------------------------------------------------------------
CORE_HEADERS = ("Month Act", "Month Bud", "Month Var", "Var %",
                "YTD Act", "YTD Bud", "YTD Var", "Var %",
                "FY Budget", "Run rate FY", "Var to FY", "F/U", "Flag")

# Widths are set here rather than per sheet, so a column that fits on one sheet
# cannot be a row of hashes on another. The full-year columns are the widest:
# an annual plan is an order of magnitude larger than a monthly one, and a
# parenthesised seven-figure variance is the longest string the layout holds.
CORE_WIDTHS = (13, 13, 13, 9, 13, 13, 13, 9, 14, 14, 14, 6, 9)

_CORE_KEYS = ("act", "bud", "var", "pct",
              "yact", "ybud", "yvar", "ypct",
              "fybud", "rr", "fyvar", "fu", "flag")


class Layout:
    """Column letters for the shared report layout.

    `label_cols` is how many description columns the sheet opens with: one on
    most sheets, three on Drivers (account, name, category). Everything after
    them is identical on every sheet and is addressed by name, never by a letter
    written out in the calling code.
    """

    #: keys whose columns hold money and take the currency format
    MONEY = ("act", "bud", "var", "yact", "ybud", "yvar", "fybud", "rr", "fyvar")
    #: keys whose columns hold percentages
    PCTS = ("pct", "ypct")

    def __init__(self, label_cols: int = 1):
        self.label_cols = label_cols
        for i, key in enumerate(_CORE_KEYS):
            setattr(self, key, get_column_letter(label_cols + 1 + i))
        self.ncols = label_cols + len(_CORE_KEYS)
        self.last_col = get_column_letter(self.ncols)

    # -- header and width rows ----------------------------------------------
    def headers(self, *label_headers: str) -> list[str]:
        return list(label_headers) + list(CORE_HEADERS)

    def widths(self, *label_widths: int) -> list[int]:
        return list(label_widths) + list(CORE_WIDTHS)

    # -- per-row cell references --------------------------------------------
    def row(self, r: int) -> dict[str, str]:
        """Every core cell reference for one row, keyed by column name."""
        return {key: f"{getattr(self, key)}{r}" for key in _CORE_KEYS}

    def money_cells(self, r: int) -> list[str]:
        return [f"{getattr(self, key)}{r}" for key in self.MONEY]

    def pct_cells(self, r: int) -> list[str]:
        return [f"{getattr(self, key)}{r}" for key in self.PCTS]

    @property
    def money_cols(self) -> list[str]:
        return [getattr(self, key) for key in self.MONEY]

    @property
    def pct_cols(self) -> list[str]:
        return [getattr(self, key) for key in self.PCTS]

    def span(self) -> list[str]:
        """Every column letter on the sheet, label columns included."""
        return [get_column_letter(c) for c in range(1, self.ncols + 1)]


def col_range(sheet: str, col: str, first: int, last: int) -> str:
    """An absolute single-column range on another sheet."""
    return f"'{sheet}'!${col}${first}:${col}${last}"


# ---------------------------------------------------------------------------
# Formula builders
# ---------------------------------------------------------------------------
def pct_f(var_cell: str, base_cell: str) -> str:
    """Variance % against a base, marked n/m only when it stops being meaningful.

    A near-zero plan produces an arithmetically correct but useless percentage,
    so those lines read n/m and are judged on the amount instead. A small but
    real base still gets a percentage.
    """
    return (f'=IF(ABS({base_cell})<{NM_ABS_FLOOR},"n/m",'
            f'IF(ABS({var_cell}/{base_cell})>{NM_RATIO_CAP},"n/m",'
            f'{var_cell}/ABS({base_cell})))')


def _material(var_cell: str, pct_cell: str, lever_eur: str, lever_pct: str) -> str:
    """TRUE when one timeframe clears both materiality floors.

    Written as nested IFs rather than AND/OR because a percentage cell can hold
    the text n/m: AND evaluates every argument, so ABS("n/m") would poison the
    whole test. IF only evaluates the branch it takes.
    """
    return (f'IF(ABS({var_cell})<{lever_eur},FALSE,'
            f'IF(NOT(ISNUMBER({pct_cell})),TRUE,ABS({pct_cell})>={lever_pct}))')


def flag_f(month_var: str, month_pct: str, ytd_var: str, ytd_pct: str,
           lever_eur: str, lever_pct: str) -> str:
    """Which timeframe is material: MONTH, YTD, BOTH or blank.

    A one-off overspend and a pattern that has been building all year need
    different answers from the reader, and a single MATERIAL badge could not
    tell them apart. Each timeframe is tested against both floors on its own
    numbers, and the flag names whichever ones clear.
    """
    m = _material(month_var, month_pct, lever_eur, lever_pct)
    y = _material(ytd_var, ytd_pct, lever_eur, lever_pct)
    return f'=IF({m},IF({y},"BOTH","MONTH"),IF({y},"YTD",""))'


def fu_f(var_cell: str, higher_is_better: bool) -> str:
    """Favourable / unfavourable badge for a variance cell."""
    return (f'=IF({var_cell}>=0,"F","U")' if higher_is_better
            else f'=IF({var_cell}<=0,"F","U")')


def run_rate_f(ytd_cell: str, months_ref: str) -> str:
    """The year to date carried across twelve months.

    Arithmetic, not a forecast: it assumes the rest of the year behaves like the
    year so far. Months elapsed is a lever cell, so a reader who disagrees with
    the assumption can change it and watch every sheet move.
    """
    return f'=IF({months_ref}=0,"",{ytd_cell}/{months_ref}*12)'


def var_to_fy_f(run_rate_cell: str, fy_budget_cell: str) -> str:
    """Where the run rate lands against the full-year plan.

    Guarded on the run rate being a number: with the months lever cleared the
    projection is blank, and blank minus a budget would report the whole annual
    plan as a variance.
    """
    return f'=IF(ISNUMBER({run_rate_cell}),{run_rate_cell}-{fy_budget_cell},"")'


class Variance:
    """Gate for the variance columns.

    A ledger extract without a plan has nothing to compare against. Writing
    `actual - 0` into the variance column would produce a full set of
    confident-looking numbers that all say "100% over budget", so when there is
    no budget the variance cells are left empty instead and the sheet says why.

    The run rate is deliberately not gated: it is built from actuals alone, so
    it stays useful on an actuals-only file.
    """

    def __init__(self, has_budget: bool):
        self.on = bool(has_budget)

    def __bool__(self) -> bool:
        return self.on

    def cell(self, formula: str):
        """The formula when a budget exists, otherwise a blank cell."""
        return formula if self.on else None

    def pct(self, var_cell: str, base_cell: str):
        return self.cell(pct_f(var_cell, base_cell))

    def flag(self, month_var: str, month_pct: str, ytd_var: str, ytd_pct: str,
             lever_eur: str, lever_pct: str):
        return self.cell(flag_f(month_var, month_pct, ytd_var, ytd_pct,
                                lever_eur, lever_pct))

    def fu(self, var_cell: str, higher_is_better: bool):
        return self.cell(fu_f(var_cell, higher_is_better))

    def var_to_fy(self, run_rate_cell: str, fy_budget_cell: str):
        return self.cell(var_to_fy_f(run_rate_cell, fy_budget_cell))


NO_BUDGET_NOTE = (
    "No budget was supplied with this file, so the budget, variance, F/U and "
    "materiality columns are intentionally left empty rather than compared "
    "against zero. The run rate is built from actuals alone, so it still "
    "projects. Upload a budget file to populate the rest."
)

SINGLE_PERIOD_NOTE = (
    "This file carries a single period, so the month and year-to-date columns "
    "describe the same postings and the full-year budget is whatever the file "
    "holds. Upload a file spanning several months to separate them."
)

LAYOUT_NOTE = (
    "Every reporting sheet carries the same columns: the month, the year to date, "
    "and the full year projected at the current run rate. F/U judges the month "
    "variance; Flag names which timeframe clears both materiality floors."
)


def meta_line(period, single_period: bool = False) -> str:
    """The right-hand line under the masthead, identical on every sheet.

    Every sheet now reports the same three horizons, so every sheet names them
    the same way. Three different phrasings were in use before - one sheet said
    "with year to date", another "YTD & FY 2025", a third listed the run rate -
    which read as three different reports.

    "YTD 2025-06" on its own would be ambiguous: it does not say whether that is
    the year through June or the month of June. Naming the window closes it.
    """
    text = str(period)
    if len(text) != 7 or text[4] != "-" or not (text[:4] + text[5:]).isdigit():
        # A file with no period column cannot be cut by time at all.
        return "All periods in the file  \u00b7  \u20ac"
    if single_period:
        return f"Month {text}  \u00b7  single period in file  \u00b7  \u20ac"
    return (f"Month {text}  \u00b7  YTD through {text}  \u00b7  FY {text[:4]}"
            f"  \u00b7  \u20ac")
