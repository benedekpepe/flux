"""
Shared Excel presentation layer: palette, type, number formats and sheet chrome.

Both packs (demo and client) render the same visual language, so the palette and
the sheet furniture live here rather than in one pack that the other imports.
Nothing in this module knows anything about finance - it only knows how a Flux
sheet is supposed to look.
"""

from __future__ import annotations
from pathlib import Path

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties


# ---------------------------------------------------------------------------
# Palette and type
# ---------------------------------------------------------------------------
FONT = "Segoe UI"

NAVY = "1B2438"
BRASS = "C6A15B"
IVORY = "F5F1E8"
BAND = "FAF8F2"
INK = "23272E"
MUTE = "8A8F98"
WHITE = "FFFFFF"

GREEN_FILL = "E7F1E8"; GREEN_INK = "2E6B3E"
RED_FILL = "F6E4E4"; RED_INK = "A63A3A"
AMBER_FILL = "FBEEDA"; AMBER_INK = "8A5A00"
# A third badge tone: the flag now names a timeframe rather than saying only
# "material", so MONTH, YTD and BOTH need to be told apart at a glance.
BLUE_FILL = "E6ECF7"; BLUE_INK = "31518C"

F_TITLE = Font(name=FONT, size=20, bold=True, color=WHITE)
F_BANDR = Font(name=FONT, size=11, color="D7DCE6")
F_META = Font(name=FONT, size=9, color=MUTE)
F_HEAD = Font(name=FONT, size=9, bold=True, color=WHITE)
F_BODY = Font(name=FONT, size=10, color=INK)
F_SMALL = Font(name=FONT, size=9, color=INK)
F_SUB = Font(name=FONT, size=10, bold=True, color=NAVY)
F_INPUT = Font(name=FONT, size=10, color="1F5FBF")
F_LABEL = Font(name=FONT, size=9, bold=True, color=MUTE)
F_KPI_LABEL = Font(name=FONT, size=9, bold=True, color=MUTE)
F_KPI_VALUE = Font(name=FONT, size=16, bold=True, color=NAVY)
F_NOTE = Font(name=FONT, size=8, italic=True, color=MUTE)
F_FU = Font(name=FONT, size=10, bold=True, color=INK)
F_FLAG = Font(name=FONT, size=9, bold=True, color=INK)

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_BRASS = PatternFill("solid", fgColor=BRASS)
FILL_HEAD = PatternFill("solid", fgColor=NAVY)
FILL_IVORY = PatternFill("solid", fgColor=IVORY)
FILL_BAND = PatternFill("solid", fgColor=BAND)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)
FILL_LEVER = PatternFill("solid", fgColor="FFF7DE")
FILL_GREEN = PatternFill("solid", fgColor=GREEN_FILL)
FILL_RED = PatternFill("solid", fgColor=RED_FILL)
FILL_AMBER = PatternFill("solid", fgColor=AMBER_FILL)
FILL_BLUE = PatternFill("solid", fgColor=BLUE_FILL)


# ---------------------------------------------------------------------------
# Number formats
# ---------------------------------------------------------------------------
# Dense report tables stay plain: the sheet header states the currency, so a
# symbol on every cell only adds noise. The symbol is used where a figure is
# read on its own - KPI cards, the materiality lever and the source amounts.
CUR = '#,##0;(#,##0);"-"'
CUR2 = '#,##0.00;(#,##0.00);"-"'
CUR_EUR = '#,##0" \u20ac";(#,##0)" \u20ac";"-"'
CUR2_EUR = '#,##0.00" \u20ac";(#,##0.00)" \u20ac";"-"'
PCT = '0.0%;(0.0%);-'
RATE = '0.0000'
KPI_DELTA = '+0.0%" vs budget";-0.0%" vs budget";0.0%" vs budget";@" vs budget"'

# Per-currency formats for local-currency amounts (the ledger mixes currencies,
# so the symbol is applied row by row from the Ccy column).
LCY_FORMATS = {
    "EUR": '#,##0.00" \u20ac";(#,##0.00)" \u20ac";"-"',
    "USD": '"$"#,##0.00;("$"#,##0.00);"-"',
    "HUF": '#,##0" Ft";(#,##0)" Ft";"-"',
}

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")


def indent(level: int) -> Alignment:
    """Left alignment with an indent, for nested rows."""
    return Alignment(horizontal="left", vertical="center", indent=level)


GOLD_SIDE = Side(style="thin", color=BRASS)
NAVY_SIDE = Side(style="medium", color=NAVY)
HEADER_BOTTOM = Border(bottom=NAVY_SIDE)
SUBTOTAL_TOP = Border(top=GOLD_SIDE)
TOTAL_TOP = Border(top=NAVY_SIDE)


def band_fill(i: int) -> PatternFill:
    """Alternating row shading."""
    return FILL_BAND if i % 2 else FILL_WHITE


def named_style(wb, name: str, *, font=None, fill=None, alignment=None,
                number_format=None) -> str:
    """Register a reusable cell style once and return its name.

    Setting `font`, `fill`, `alignment` and `number_format` separately makes
    openpyxl re-hash the style on each assignment, which on a sheet of five
    thousand posting lines is the single largest cost in building the pack. A
    named style collapses those four writes into one lookup.
    """
    from openpyxl.styles import NamedStyle

    if name in wb.named_styles:
        return name
    style = NamedStyle(name=name)
    if font is not None:
        style.font = font
    if fill is not None:
        style.fill = fill
    if alignment is not None:
        style.alignment = alignment
    if number_format is not None:
        style.number_format = number_format
    wb.add_named_style(style)
    return name


def wrapped_height(text, col_width: int, base: int = 26, per_line: int = 13) -> int:
    """Row height that fits `text` wrapped inside a column of `col_width`.

    openpyxl cannot measure text, and Excel only auto-fits a wrapped row when no
    height is set - which these sheets do set, to keep the table rhythm even.
    So the height is estimated instead: a client chart of accounts with long
    names would otherwise have its commentary silently clipped.
    """
    if not text:
        return base
    lines = max(1, -(-len(str(text)) // max(1, col_width - 2)))
    return max(base, 6 + per_line * lines)


# ---------------------------------------------------------------------------
# Sheet chrome
# ---------------------------------------------------------------------------
def hide_grid(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = NAVY
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True


def title_band(ws, subtitle, meta_right, last_col, entity="Demo Company Ltd") -> None:
    """Navy masthead, entity/meta line and the brass rule beneath it.

    The wordmark takes two columns rather than four: the subtitle is
    right-aligned inside its merged range, so on a narrow sheet a wide
    wordmark block would push it out of the range and clip it on the left.
    """
    ws.merge_cells("A1:B1"); ws["A1"] = "FLUX"; ws["A1"].font = F_TITLE
    ws["A1"].alignment = LEFT
    ws.merge_cells(f"C1:{last_col}1"); ws["C1"] = subtitle; ws["C1"].font = F_BANDR
    ws["C1"].alignment = RIGHT
    for c in range(1, ws[f"{last_col}1"].column + 1):
        ws.cell(row=1, column=c).fill = FILL_NAVY
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:B2"); ws["A2"] = entity; ws["A2"].font = F_META
    ws["A2"].alignment = LEFT
    ws.merge_cells(f"C2:{last_col}2"); ws["C2"] = meta_right; ws["C2"].font = F_META
    ws["C2"].alignment = RIGHT
    ws.row_dimensions[2].height = 16

    ws.merge_cells(f"A3:{last_col}3")
    for c in range(1, ws[f"{last_col}3"].column + 1):
        ws.cell(row=3, column=c).fill = FILL_BRASS
    ws.row_dimensions[3].height = 3


def headers(ws, hrow, labels, center_from=2, center_to=None) -> None:
    center_to = center_to if center_to is not None else len(labels)
    for c, h in enumerate(labels, start=1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        cell.alignment = CENTER if center_from <= c <= center_to else LEFT
        cell.border = HEADER_BOTTOM
    ws.row_dimensions[hrow].height = 22


def widths(ws, values, start=1) -> None:
    for c, w in enumerate(values, start=start):
        ws.column_dimensions[get_column_letter(c)].width = w


def fit_text_columns(ws, letters, first_row, last_row, *, pad=2.0,
                     max_width=40.0) -> None:
    """Widen text columns to the longest label they actually hold.

    Excel clips a string at the cell edge whenever the neighbouring cell is
    occupied, so a name column sized by guesswork silently truncates the longest
    account names. Measuring the content is the only way to be right for a chart
    of accounts nobody has seen yet; the cap stops one outlier stretching a
    column across the sheet.
    """
    for letter in letters:
        longest = 0
        for row in range(first_row, last_row + 1):
            value = ws[f"{letter}{row}"].value
            if isinstance(value, str) and not value.startswith("="):
                longest = max(longest, len(value))
        if longest:
            current = ws.column_dimensions[letter].width or 8.43
            ws.column_dimensions[letter].width = min(
                max_width, max(current, longest + pad))


def outline(ws, r1, c1, r2, c2, side=GOLD_SIDE) -> None:
    """Draw a border around a rectangular range without clearing inner edges."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            cell.border = Border(
                top=side if r == r1 else b.top,
                bottom=side if r == r2 else b.bottom,
                left=side if c == c1 else b.left,
                right=side if c == c2 else b.right,
            )


# The flag now names a timeframe rather than asserting a single verdict, so each
# value gets its own weight: a month-only movement is the new news, a
# year-to-date-only one is a pattern the month happens not to show, and BOTH is
# both at once - the line a reader should look at first.
FLAG_BADGES = {
    "MONTH": (FILL_AMBER, AMBER_INK),
    "YTD": (FILL_BLUE, BLUE_INK),
    "BOTH": (FILL_RED, RED_INK),
}


def lever(ws, label_range: str, label: str, cell: str, value,
          number_format: str) -> None:
    """One editable assumption: a caption and a boxed input cell.

    Three of these now sit on the P&L - two materiality floors and months
    elapsed - and every other sheet points at them, so they are written from one
    place rather than described twice in two packs.
    """
    ws.merge_cells(label_range)
    head = label_range.split(":")[0]
    ws[head] = label; ws[head].font = F_LABEL; ws[head].alignment = LEFT
    ws[cell] = value
    ws[cell].font = F_INPUT
    ws[cell].number_format = number_format
    ws[cell].fill = FILL_LEVER
    ws[cell].alignment = CENTER
    target = ws[cell]
    outline(ws, target.row, target.column, target.row, target.column, GOLD_SIDE)


def badge_cf(ws, rng_fu, rng_flag) -> None:
    """Green/red F and U badges; amber MONTH, blue YTD and red BOTH flags."""
    if rng_fu:
        ws.conditional_formatting.add(rng_fu, CellIsRule(
            operator="equal", formula=['"F"'], fill=FILL_GREEN,
            font=Font(name=FONT, size=10, bold=True, color=GREEN_INK)))
        ws.conditional_formatting.add(rng_fu, CellIsRule(
            operator="equal", formula=['"U"'], fill=FILL_RED,
            font=Font(name=FONT, size=10, bold=True, color=RED_INK)))
    if rng_flag:
        # BOTH first: Excel applies matching rules in order, and an "equal"
        # test is exact, but keeping the strongest badge at the top means a
        # later change to a prefix match cannot silently downgrade it.
        for value in ("BOTH", "MONTH", "YTD"):
            fill, ink = FLAG_BADGES[value]
            ws.conditional_formatting.add(rng_flag, CellIsRule(
                operator="equal", formula=[f'"{value}"'], fill=fill,
                font=Font(name=FONT, size=9, bold=True, color=ink)))


def variance_cf(ws, rng, fu_ref) -> None:
    """Colour a variance range from the favourable/unfavourable verdict.

    The sign alone cannot carry the meaning: revenue below plan and cost above
    plan are both bad news with opposite signs. Driving the colour from the F/U
    cell keeps the numbers and the badge telling the same story, and matches how
    the deck renders the same figures.

    Only the colour is set, so the weight of a subtotal row survives.
    """
    from openpyxl.formatting.rule import FormulaRule

    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'{fu_ref}="F"'], font=Font(color=GREEN_INK)))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'{fu_ref}="U"'], font=Font(color=RED_INK)))


def spend_variance_cf(ws, rng) -> None:
    """Colour a variance range where every line is a cost.

    Lower is better on all of them, so the sign is enough.
    """
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThan", formula=["0"], font=Font(color=RED_INK)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0"], font=Font(color=GREEN_INK)))


def signed_variance_cf(ws, ranges, higher_is_better: bool) -> None:
    """Colour a variance range whose favourable direction is known at build time.

    The sheet carries one F/U column but three variances - month, year to date
    and against the full-year plan - and they do not always share a sign. Only
    the month variance can be driven from the badge; the other two are coloured
    from the direction the line actually has, which is known when the row is
    written.

    `ranges` is a whitespace-separated sqref, so the rows sharing a direction on
    a mixed sheet are collected into one rule instead of one rule per row.
    """
    if not ranges:
        return
    good, bad = ("greaterThan", "lessThan") if higher_is_better else ("lessThan", "greaterThan")
    ws.conditional_formatting.add(ranges, CellIsRule(
        operator=good, formula=["0"], font=Font(color=GREEN_INK)))
    ws.conditional_formatting.add(ranges, CellIsRule(
        operator=bad, formula=["0"], font=Font(color=RED_INK)))


def sqref(cols, rows) -> str:
    """A whitespace-separated range covering `cols` on each row in `rows`.

    Consecutive rows are merged into one range so the sqref stays short: Excel
    accepts a list of single cells, but a rule written that way is unreadable in
    the manager dialog and slow to open on a sheet of any size.
    """
    rows = sorted(set(rows))
    if not rows or not cols:
        return ""
    blocks, start, prev = [], rows[0], rows[0]
    for row in rows[1:]:
        if row == prev + 1:
            prev = row
            continue
        blocks.append((start, prev)); start = prev = row
    blocks.append((start, prev))
    first_col, last_col = cols[0], cols[-1]
    return " ".join(f"{first_col}{a}:{last_col}{b}" for a, b in blocks)


def note(ws, row, text) -> None:
    ws.cell(row=row, column=1, value=text).font = F_NOTE


# ---------------------------------------------------------------------------
# Excel's green corner markers
# ---------------------------------------------------------------------------
def quiet_indicators(ws, first_row, last_row, last_col="Z") -> None:
    """Record the table range whose Excel warning markers should be hidden.

    Subtotal rows legitimately use different formulas from the detail rows above
    them, which Excel flags as an inconsistency with a green corner marker. The
    marks are noise on a finished report. openpyxl has no writer for the
    element, so the range is remembered here and injected after saving.
    """
    ws._flux_quiet_range = f"A{first_row}:{last_col}{max(first_row, last_row)}"


def collect_quiet_ranges(wb) -> dict[str, str]:
    return {name: getattr(wb[name], "_flux_quiet_range", None)
            for name in wb.sheetnames
            if getattr(wb[name], "_flux_quiet_range", None)}


def suppress_error_indicators(path: Path, ranges: dict[str, str]) -> None:
    """Inject <ignoredErrors> into a saved workbook.

    Written straight into the sheet XML because openpyxl has no writer for it.
    The element must sit after the page/break elements and before any drawing,
    table or extension list, per the sheet schema.
    """
    import re
    import shutil
    import zipfile
    from xml.sax.saxutils import escape

    path = Path(path)
    src = zipfile.ZipFile(path)
    name_by_sheet = {}
    wb_xml = src.read("xl/workbook.xml").decode("utf-8")
    order = re.findall(r'<sheet[^>]*name="([^"]+)"', wb_xml)
    for i, sheet_name in enumerate(order, start=1):
        name_by_sheet[f"xl/worksheets/sheet{i}.xml"] = sheet_name

    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            sheet_name = name_by_sheet.get(item.filename)
            rng = ranges.get(sheet_name) if sheet_name else None
            if rng:
                xml = data.decode("utf-8")
                block = (f'<ignoredErrors><ignoredError sqref="{escape(rng)}" '
                         f'formula="1" formulaRange="1" numberStoredAsText="1" '
                         f'emptyCellReference="1" evalError="1"/></ignoredErrors>')
                anchor = next((t for t in ("<drawing", "<legacyDrawing", "<tableParts",
                                           "<extLst") if t in xml), None)
                if anchor:
                    xml = xml.replace(anchor, block + anchor, 1)
                else:
                    xml = xml.replace("</worksheet>", block + "</worksheet>", 1)
                data = xml.encode("utf-8")
            dst.writestr(item, data)
    src.close()
    shutil.move(str(tmp), str(path))
