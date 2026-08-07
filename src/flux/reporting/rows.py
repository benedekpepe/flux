"""
The shared report row.

Both packs write the same thirteen columns on every reporting sheet. A sheet
knows what it is cutting the ledger by and therefore supplies five figures - the
month actual and budget, the year-to-date pair, and the full-year plan. Nothing
after that differs: the variances, the run-rate projection, the F/U badge, the
materiality flag and their formats follow from those five.

Keeping that tail in one place is the point. It lived twice before, once per
pack, and the two copies drifted: different column orders, one sheet with a
projection and one without, a materiality flag that meant different things in
different places. This module is what stops that happening again.

It sits between `formulas` (which knows the arithmetic and no styling) and
`styling` (which knows the look and no finance), because the report row needs
both and neither should import the other.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter

from .formulas import (Layout, PL_E, PL_P, PL_M, flag_f, fu_f, pct_f,
                       run_rate_f, var_to_fy_f)
from .styling import (CENTER, CUR, CUR_EUR, F_BODY, F_ECHO, F_ECHO_LABEL,
                      F_FLAG, F_FU, F_NOTE, F_SUB, LEFT, PCT, RIGHT,
                      badge_cf, signed_variance_cf, sqref, variance_cf)

# Keys of the five figures a sheet has to supply itself.
BASE_KEYS = ("act", "bud", "yact", "ybud", "fybud")


def write_tail(ws, L: Layout, r: int, *, higher: bool, bold: bool = False,
               var=None, lev_e: str = PL_E, lev_p: str = PL_P,
               months: str = PL_M) -> None:
    """Fill in and format everything that follows from the five base figures.

    `var` is the no-budget gate: on an actuals-only file the variance, F/U and
    flag cells are left empty rather than measured against zero. The run rate is
    written either way, because it is built from actuals alone.
    """
    c = L.row(r)
    gate = (lambda f: f) if var is None else var.cell

    ws[c["var"]] = gate(f'={c["act"]}-{c["bud"]}')
    ws[c["pct"]] = gate(pct_f(c["var"], c["bud"]))
    ws[c["yvar"]] = gate(f'={c["yact"]}-{c["ybud"]}')
    ws[c["ypct"]] = gate(pct_f(c["yvar"], c["ybud"]))
    ws[c["rr"]] = run_rate_f(c["yact"], months)
    ws[c["fyvar"]] = gate(var_to_fy_f(c["rr"], c["fybud"]))
    ws[c["fypct"]] = gate(pct_f(c["fyvar"], c["fybud"]))
    ws[c["fu"]] = gate(fu_f(c["var"], higher))
    ws[c["flag"]] = gate(flag_f(c["var"], c["pct"], c["yvar"], c["ypct"],
                                lev_e, lev_p))

    body = F_SUB if bold else F_BODY
    for ref in L.money_cells(r):
        cell = ws[ref]; cell.number_format = CUR; cell.font = body
        cell.alignment = RIGHT
    for ref in L.pct_cells(r):
        cell = ws[ref]; cell.number_format = PCT; cell.font = body
        cell.alignment = RIGHT
    ws[c["fu"]].alignment = CENTER; ws[c["fu"]].font = F_FU
    ws[c["flag"]].alignment = CENTER; ws[c["flag"]].font = F_FLAG


def write_sum_tail(ws, L: Layout, r: int, rows, *, higher: bool, var=None,
                   lev_e: str = PL_E, lev_p: str = PL_P,
                   months: str = PL_M) -> None:
    """A subtotal row: add the base figures over `rows`, then the usual tail.

    The subtotal adds the five bases and re-derives everything else rather than
    adding the variances, so a subtotal percentage is the percentage of the
    subtotal and not the sum of percentages.
    """
    gate = (lambda f: f) if var is None else var.cell
    for key in BASE_KEYS:
        col = getattr(L, key)
        formula = "=" + "+".join(f"{col}{x}" for x in rows)
        # Actuals are always known; only the plan-derived columns are gated.
        ws[f"{col}{r}"] = formula if key in ("act", "yact") else gate(formula)
    write_tail(ws, L, r, higher=higher, bold=True, var=var,
               lev_e=lev_e, lev_p=lev_p, months=months)


def lever_echo(ws, L: Layout) -> None:
    """Show the P&L's lever values on a sheet that does not own them.

    A reader looking at BOTH on the expense report cannot see what threshold
    produced it without leaving the sheet, and the same goes for the month count
    behind every run rate. So each sheet repeats the three numbers on the row
    under the masthead.

    Deliberately read-only, and styled to look it: no input colour, no box. Two
    editable copies of one assumption is two sources of truth, which is the
    problem the lever cells exist to solve. These are formulas pointing at the
    P&L, so they follow it.
    """
    for label_range, label, cell, ref, fmt in (
            ("A4:B4", "Materiality floor (\u20ac)", "C4", PL_E, CUR_EUR),
            ("E4:F4", "Materiality floor (%)", "G4", PL_P, PCT),
            ("I4:J4", "Months elapsed", "K4", PL_M, "0")):
        ws.merge_cells(label_range)
        head = label_range.split(":")[0]
        ws[head] = label; ws[head].font = F_ECHO_LABEL; ws[head].alignment = LEFT
        ws[cell] = f"={ref}"
        ws[cell].font = F_ECHO; ws[cell].number_format = fmt
        ws[cell].alignment = CENTER
    note_col = get_column_letter(max(12, L.ncols - 3))
    ws.merge_cells(f"{note_col}4:{L.last_col}4")
    ws[f"{note_col}4"] = "set on the P&L Report"
    ws[f"{note_col}4"].font = F_NOTE
    ws[f"{note_col}4"].alignment = RIGHT
    ws.row_dimensions[4].height = 15


def report_cf(ws, L: Layout, first: int, last: int, rows_higher, rows_lower,
              *, var=None) -> None:
    """Badges, and the three variance blocks coloured by their own direction.

    The month variance is coloured from the F/U cell beside it, so the number
    and the badge cannot disagree. The year-to-date and full-year variances get
    their own rules: they do not always share a sign with the month, and one
    badge cannot speak for three timeframes. Rows sharing a direction are
    collected into a single rule rather than one rule per row.

    Ranges run through the total row, not the last detail row: a total carries
    the same verdict as the lines above it, and an off-by-one there leaves it
    rendering black while everything above it is green or red.
    """
    badge_cf(ws, f"{L.fu}{first}:{L.fu}{last}", f"{L.flag}{first}:{L.flag}{last}")
    if var is not None and not var:
        return
    variance_cf(ws, f"{L.var}{first}:{L.pct}{last}", f"${L.fu}{first}")
    for rows, higher in ((rows_higher, True), (rows_lower, False)):
        signed_variance_cf(ws, sqref([L.yvar, L.ypct], rows), higher)
        signed_variance_cf(ws, sqref([L.fyvar, L.fypct], rows), higher)
