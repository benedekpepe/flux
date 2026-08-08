"""
Demo pack: the full multi-entity showcase, built from generated data.

Inputs (how a real finance function holds data):
  - GL Transactions : year-to-date transaction-level actuals, multi-entity,
                      multi-currency (with FX to EUR), tagged with cost centre,
                      expense type, posting date and source journal.
  - Budget          : full-year summarised plan (and prior year) by entity x
                      cost centre x account x period.

Every reporting sheet is a live formula view over those two inputs, and all
views reconcile to the source. They also share one set of columns - the month,
the year to date, the full-year plan and the run-rate projection against it,
then F/U and the materiality flag - so the sheets differ in what they cut the
ledger by, never in how they report it:
  - P&L Report     consolidated P&L, KPI cards, per-line commentary, and the
                   three lever cells the rest of the pack points at
  - Expense Report natural view: expense type
  - By Entity      consolidation: net income per legal entity
  - Departments & CCs  functional view: department roll-up with cost centres
  - Drivers        account-level movers with data bars

Output: output/flux_demo_pack.xlsx
"""

from __future__ import annotations
from pathlib import Path

import pandas as pd

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..coa import (PNL_STRUCTURE, CATEGORY_FAVOURABLE, FAV_HIGHER, COST_CENTRES,
                   SPEND_DEPARTMENTS, DEPARTMENT_COST_CENTRES, EXPENSE_GROUPS)
from ..analysis import NO_CAUSE_NOTE
from ..commentary import _join_names, _mag, _pct
from ..commentary import line_comments, rollup_comments, total_comment
from ..engine import build_report
from .styling import (
    F_HEAD, F_BODY, F_SMALL, F_SUB, F_INPUT, F_KPI_LABEL, F_KPI_VALUE, F_NOTE,
    F_META, F_KPI_DELTA,
    FILL_HEAD, FILL_IVORY, FILL_BAND, FILL_WHITE,
    CUR2, CUR_EUR, CUR2_EUR, PCT, RATE, KPI_DELTA, KPI_MONTH, KPI_MONTH_PCT,
    LCY_FORMATS,
    CENTER, LEFT, RIGHT, WRAP, indent, band_fill,
    GOLD_SIDE, HEADER_BOTTOM, SUBTOTAL_TOP, TOTAL_TOP,
    F_LABEL,
    hide_grid, title_band, headers, widths, outline, lever,
    wrapped_height, fit_text_columns,
    signed_variance_cf,
    named_style,
    quiet_indicators, collect_quiet_ranges, suppress_error_indicators,
)
# The columns every reporting sheet shares are written by `rows`, so a sheet
# here only supplies the five figures that depend on how it cuts the ledger.
from .rows import (lever_echo as _lever_echo, report_cf as _report_cf,
                   write_sum_tail as _sum_tail, write_tail as _tail)
from .formulas import (
    Layout,
    LEVER_EUR, LEVER_PCT, LEVER_MONTHS, LEV_E, LEV_P, LEV_M,
    DEFAULT_ABS_THRESHOLD, DEFAULT_PCT_THRESHOLD, meta_line,
)


# Width of the commentary column, shared by the column setup and the row-height
# estimate so the two cannot drift apart.
COMMENT_WIDTH = 58

# ---- source-sheet column letters ------------------------------------------
# ---- source-sheet layout -----------------------------------------------------
# (key, header, width, format) - column letters are derived, never hand-counted.
GL_SPEC = [
    ("doc_no", "Document", 12, "text"),
    ("doc_type", "Type", 6, "text"),
    ("doc_type_text", "Document type", 18, "text"),
    ("fiscal_year", "FY", 7, "int"),
    ("period", "Period", 9, "text"),
    ("period_no", "Per. key", 9, "int"),
    ("doc_date", "Doc date", 11, "text"),
    ("posting_date", "Posting date", 12, "text"),
    ("entry_date", "Entry date", 11, "text"),
    ("entry_time", "Time", 7, "text"),
    ("reference", "Reference", 14, "text"),
    ("header_text", "Header text", 24, "text"),
    ("created_by", "Created by", 12, "text"),
    ("reversed", "Rev.", 6, "text"),
    ("entity", "Entity", 12, "text"),
    ("company_code", "Co. code", 9, "text"),
    ("region", "Region", 10, "text"),
    ("department", "Department", 13, "text"),
    ("cost_centre", "Cost centre", 12, "text"),
    ("profit_centre", "Profit centre", 13, "text"),
    ("line_no", "Line", 6, "int"),
    ("account_code", "Account", 9, "text"),
    ("account_name", "Account name", 26, "text"),
    ("category", "Category", 11, "text"),
    ("expense_type", "Expense type", 22, "text"),
    ("group", "P&L group", 20, "text"),
    ("dc_indicator", "D/C", 6, "text"),
    ("currency", "Ccy", 6, "text"),
    ("amount_lcy", "Amount (local)", 15, "lcy"),
    ("fx_rate", "FX", 9, "rate"),
    ("amount_eur", "Amount (\u20ac)", 15, "cur2"),
    ("tax_code", "Tax", 6, "text"),
    ("partner", "Partner", 24, "text"),
    ("assignment", "Assignment", 18, "text"),
    ("line_text", "Item text", 44, "text"),
    ("source", "Source", 9, "text"),
]

BUD_SPEC = [
    ("period", "Period", 9, "text"),
    ("period_no", "Per. key", 9, "int"),
    ("version", "Version", 12, "text"),
    ("entity", "Entity", 13, "text"),
    ("region", "Region", 10, "text"),
    ("currency", "Ccy", 6, "text"),
    ("department", "Department", 13, "text"),
    ("cost_centre", "Cost centre", 12, "text"),
    ("account_code", "Account", 9, "text"),
    ("account_name", "Account name", 26, "text"),
    ("category", "Category", 11, "text"),
    ("expense_type", "Expense type", 22, "text"),
    ("group", "P&L group", 20, "text"),
    ("owner", "Budget owner", 14, "text"),
    ("budget_eur", "Budget", 14, "cur"),
    ("prior_eur", "Prior Yr Act", 15, "cur"),
]


def _letters(spec):
    return {key: get_column_letter(i) for i, (key, *_rest) in enumerate(spec, start=1)}


GL_COL = _letters(GL_SPEC)
BUD_COL = _letters(BUD_SPEC)

GL_PERIOD, GL_ENT, GL_DEPT = GL_COL["period"], GL_COL["entity"], GL_COL["department"]
GL_PERNO = GL_COL["period_no"]
GL_CC, GL_CODE, GL_CAT = GL_COL["cost_centre"], GL_COL["account_code"], GL_COL["category"]
GL_EXP, GL_EUR = GL_COL["expense_type"], GL_COL["amount_eur"]
BUD_PERIOD, BUD_ENT, BUD_DEPT = BUD_COL["period"], BUD_COL["entity"], BUD_COL["department"]
BUD_PERNO = BUD_COL["period_no"]
BUD_CC, BUD_CODE, BUD_CAT = BUD_COL["cost_centre"], BUD_COL["account_code"], BUD_COL["category"]
BUD_EXP, BUD_BUD, BUD_PRIOR = BUD_COL["expense_type"], BUD_COL["budget_eur"], BUD_COL["prior_eur"]

# ===========================================================================
# Input sheets
# ===========================================================================
def _write_source_sheet(ws, df, spec, title, meta, footnote) -> tuple[int, int]:
    """Write an input sheet from a column spec; returns (first_row, last_row)."""
    hide_grid(ws)
    last_col = get_column_letter(len(spec))
    title_band(ws, title, meta, last_col)

    numeric = {"cur", "cur2", "lcy", "rate", "int"}
    hrow = 5
    for i, (key, header, width, fmt) in enumerate(spec, start=1):
        cell = ws.cell(row=hrow, column=i, value=header)
        cell.font = F_HEAD; cell.fill = FILL_HEAD
        cell.alignment = CENTER if fmt in numeric else LEFT
        cell.border = HEADER_BOTTOM
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[hrow].height = 22

    fmt_map = {"cur": CUR_EUR, "cur2": CUR2_EUR, "rate": RATE, "int": "0"}
    wb = ws.parent

    # Resolved names are memoised: openpyxl's `name in wb.named_styles` rebuilds
    # the name list on every call, which at a quarter of a million cells costs
    # more than the styling it was meant to save.
    style_cache: dict[tuple, str] = {}

    def style_for(fmt, banded, ccy="EUR"):
        """One registered style per (format, band, currency) combination.

        Roughly a dozen styles cover a sheet of any length, so the per-cell work
        drops from four style writes to one dictionary hit.
        """
        cached = style_cache.get((fmt, banded, ccy))
        if cached is not None:
            return cached
        suffix = "b" if banded else "w"
        fill = FILL_BAND if banded else FILL_WHITE
        if fmt == "lcy":
            name = named_style(wb, f"flux_lcy_{ccy}_{suffix}", font=F_INPUT,
                               fill=fill, alignment=RIGHT,
                               number_format=LCY_FORMATS.get(ccy, CUR2))
        elif fmt in ("cur", "cur2"):
            name = named_style(wb, f"flux_{fmt}_{suffix}", font=F_INPUT, fill=fill,
                               alignment=RIGHT, number_format=fmt_map[fmt])
        elif fmt in ("rate", "int"):
            name = named_style(wb, f"flux_{fmt}_{suffix}", font=F_SMALL, fill=fill,
                               alignment=RIGHT, number_format=fmt_map[fmt])
        else:
            name = named_style(wb, f"flux_text_{suffix}", font=F_SMALL, fill=fill,
                               alignment=LEFT)
        style_cache[(fmt, banded, ccy)] = name
        return name

    first = hrow + 1
    # Plain dicts: `row.get(key)` on a Series goes through pandas indexing on
    # every cell, and there are a quarter of a million of them here.
    for i, row in enumerate(df.to_dict("records")):
        r = first + i
        banded = bool(i % 2)
        for c, (key, _h, _w, fmt) in enumerate(spec, start=1):
            v = row.get(key, "")
            if fmt in numeric and v != "":
                v = float(v) if fmt != "int" else int(v)
            cell = ws.cell(row=r, column=c, value=v)
            cell.style = style_for(fmt, banded, str(row.get("currency", "EUR")))
        ws.row_dimensions[r].height = 14
    last = first + len(df) - 1

    text_cols = [get_column_letter(i) for i, (_k, _h, _w, fmt) in enumerate(spec, start=1)
                 if fmt not in numeric]
    fit_text_columns(ws, text_cols, first, last)

    ws.cell(row=last + 2, column=1, value=footnote).font = F_NOTE
    quiet_indicators(ws, 4, last)
    # Input sheets are wide too: hold the first column as well as the header.
    ws.freeze_panes = f"B{first}"
    return first, last


def _write_gl_transactions(ws, txns, period) -> tuple[int, int]:
    return _write_source_sheet(
        ws, txns, GL_SPEC,
        "GL Transactions · year-to-date posting lines (multi-entity, multi-currency)",
        f"Year to date through {period}  ·  reporting currency \u20ac",
        "Blue cells are input amounts. Amount (EUR) = Amount (LCY) x FX. "
        "Revenue is credited (H), costs are debited (S). Only account, category, "
        "cost centre, period and the EUR amount drive the reports; the remaining "
        "columns are carried for traceability.",
    )


def _write_budget(ws, bud, period) -> tuple[int, int]:
    return _write_source_sheet(
        ws, bud, BUD_SPEC,
        "Budget · full-year plan by entity x department x cost centre x account",
        "Full year 2025  ·  \u20ac",
        "Blue cells are the plan (budget) and prior-year actual figures.",
    )


# Cost centres are keyed by code on the sheet but read by name in prose.
TOTAL_KEY = "\u0000total"
CC_NAMES = {code: name for _dept, code, name in COST_CENTRES}

# The expense groups are a constant, not a column, so a roll-up commentary over
# them needs the mapping materialised onto the frame first.
_GROUP_OF_TYPE = {t: name for name, types in EXPENSE_GROUPS for t in types}


def _with_group(detail):
    """The detail frame with its expense group attached."""
    out = detail.copy()
    out["expense_group"] = out["expense_type"].map(_GROUP_OF_TYPE).fillna("Other expense types")
    return out


def _spend_detail(detail):
    """The detail frame with revenue dropped.

    The expense and department sheets are spend views: they exclude revenue from
    every figure on them. A commentary built on the unfiltered frame agrees with
    the individual rows by luck - the departments that book revenue are not on
    the sheet - and then contradicts the grand total, which sums everything.
    """
    return detail[detail["category"] != "Revenue"]


def _net_detail(detail):
    """The detail frame signed so that a total reads as net income.

    The By Entity sheet reports revenue less spend, so a commentary built on the
    raw frame would add the two together and cheerfully call an overspending
    entity favourable - a comment contradicting the row it sits beside.
    """
    out = detail.copy()
    sign = out["category"].eq("Revenue").map({True: 1.0, False: -1.0})
    for col in ("actual", "budget"):
        out[col] = out[col] * sign
    return out


def _ytd_detail(txns, bud, perno):
    """Year-to-date actuals and plan at the same grain as `monthly_detail`."""
    keys = ["entity", "department", "cost_centre", "account_code",
            "account_name", "category", "expense_type"]
    act = (txns[txns["period_no"] <= perno].groupby(keys, as_index=False)["amount_eur"]
           .sum().rename(columns={"amount_eur": "actual"}))
    plan = (bud[bud["period_no"] <= perno].groupby(keys, as_index=False)["budget_eur"]
            .sum().rename(columns={"budget_eur": "budget"}))
    return act.merge(plan, on=keys, how="outer").fillna(0.0)


def _write_analysis(ws, L, first_row, blocks, com_width) -> int:
    """The analysis section under a report table.

    Written only for the lines the flag already picked out: a block under every
    row would bury the two that matter. Each one names the line, then what the
    figures on the sheet imply about it - never why it moved, which the ledger
    does not record.
    """
    if not blocks:
        return first_row
    r = first_row
    ws.cell(row=r, column=1, value="Analysis \u00b7 what the numbers point at").font = F_SUB
    for col in L.span():
        ws[f"{col}{r}"].border = TOTAL_TOP
    ws.row_dimensions[r].height = 22
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=L.ncols)
    ws.cell(row=r, column=1, value=NO_CAUSE_NOTE).font = F_NOTE
    ws.cell(row=r, column=1).alignment = WRAP
    ws.row_dimensions[r].height = wrapped_height(NO_CAUSE_NOTE, com_width * 2)
    r += 2

    # EBIT and net income reach the same findings - every category's movement
    # lands on both - so printing both in full is one thought written twice.
    # The second gets a cross-reference instead, which is shorter and truer.
    seen = {}
    for label, flag, items in blocks:
        key = tuple(f.text for f in items if f.heading != "Full year")
        if key in seen:
            ws.cell(row=r, column=1, value=f"{label}   \u00b7   {flag}" if flag else label).font = F_SUB
            ws.cell(row=r, column=1).alignment = LEFT
            for col in L.span():
                ws[f"{col}{r}"].fill = FILL_IVORY
            ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1,
                           end_column=L.ncols)
            note = (f"The same picture as {seen[key]} above: the movement, its "
                    "persistence and the question it raises are identical, "
                    "because the lines below feed both.")
            ws.cell(row=r + 1, column=1, value=note).font = F_BODY
            ws.cell(row=r + 1, column=1).alignment = WRAP
            ws.row_dimensions[r + 1].height = wrapped_height(note, com_width * 2)
            r += 3
            continue
        seen[key] = label
        title = f"{label}   \u00b7   {flag}" if flag else label
        ws.cell(row=r, column=1, value=title).font = F_SUB
        ws.cell(row=r, column=1).alignment = LEFT
        for col in L.span():
            ws[f"{col}{r}"].fill = FILL_IVORY
        ws.row_dimensions[r].height = 20
        r += 1
        for finding in items:
            head = ws.cell(row=r, column=1, value=finding.heading)
            head.font = F_LABEL; head.alignment = indent(1)
            ws.merge_cells(start_row=r, start_column=2, end_row=r,
                           end_column=L.ncols)
            body = ws.cell(row=r, column=2, value=finding.text)
            body.font = F_BODY; body.alignment = WRAP
            ws.row_dimensions[r].height = wrapped_height(finding.text, com_width * 2)
            r += 1
        r += 1
    return r


# ---------------------------------------------------------------------------
# Analysis inputs
# ---------------------------------------------------------------------------
def _history(txns, bud, mask_t, mask_b, perno):
    """Actual and plan per period, up to the reporting month."""
    act = (txns[mask_t & (txns["period_no"] <= perno)]
           .groupby(["period", "period_no"], as_index=False)["amount_eur"].sum()
           .rename(columns={"amount_eur": "actual"}))
    plan = (bud[mask_b & (bud["period_no"] <= perno)]
            .groupby(["period", "period_no"], as_index=False)["budget_eur"].sum()
            .rename(columns={"budget_eur": "budget"}))
    return act.merge(plan, on=["period", "period_no"], how="outer").fillna(0.0)


def _flag_of(month_var, month_pct, ytd_var, ytd_pct):
    """The same MONTH / YTD / BOTH verdict the sheet's Flag column reaches."""
    def material(var, pct):
        return (abs(var) >= DEFAULT_ABS_THRESHOLD
                and (pct is None or abs(pct) >= DEFAULT_PCT_THRESHOLD))
    m, y = material(month_var, month_pct), material(ytd_var, ytd_pct)
    return "BOTH" if m and y else "MONTH" if m else "YTD" if y else ""


def _pct_or_none(var, base):
    return None if abs(base) < 100 else var / abs(base)


def _computed_block(out, line, line_var, ytd_line_var, agg, ytd_agg, txns, bud,
                    perno, months, higher):
    """Analysis for a subtotal, attributed to the lines that make it up."""
    from ..analysis import findings

    parts, feed = [], []
    stack = list(line.components)
    structure = {l.label: l for l in PNL_STRUCTURE}
    while stack:
        sign, ref = stack.pop()
        node = structure[ref]
        if node.kind == "category":
            # A cost below a profit line pushes it the other way, so the sign of
            # its contribution is the opposite of the sign of its variance.
            direction = 1 if sign == "+" else -1
            # Cumulative, like every other finding in the block.
            parts.append({"account_name": node.label,
                          "var_bud": direction * ytd_line_var[node.label]})
            feed.append((node.category, direction))
        else:
            flip = 1 if sign == "+" else -1
            stack += [(("+" if (s == "+") == (flip == 1) else "-"), r)
                      for s, r in node.components]
    m_var = line_var[line.label]
    y_var = _signed_total(ytd_agg, line, structure)
    flag = _flag_of(m_var, _pct_or_none(m_var, _signed_base(agg, line, structure)),
                    y_var, _pct_or_none(y_var, _signed_base(ytd_agg, line, structure)))
    if not flag:
        return
    # Every figure below the flag has to carry the line's own signs: a subtotal
    # summed unsigned is revenue plus costs, which reported a €42m plan for a
    # €3.6m EBIT line and called a losing month favourable.
    signs = dict(feed)
    t_sign = txns["category"].map(signs)
    b_sign = bud["category"].map(signs)
    t = txns.assign(_v=txns["amount_eur"] * t_sign).dropna(subset=["_v"])
    b = bud.assign(_v=bud["budget_eur"] * b_sign).dropna(subset=["_v"])
    hist = (t[t["period_no"] <= perno]
            .groupby(["period", "period_no"], as_index=False)["_v"].sum()
            .rename(columns={"_v": "actual"})
            .merge(b[b["period_no"] <= perno]
                   .groupby(["period", "period_no"], as_index=False)["_v"].sum()
                   .rename(columns={"_v": "budget"}),
                   on=["period", "period_no"], how="outer").fillna(0.0))
    ytd_actual = t.loc[t["period_no"] <= perno, "_v"].sum()
    out.append((line.label, flag, findings(
        drivers=pd.DataFrame(parts), total_var=y_var, history=hist,
        run_rate=ytd_actual / months * 12 if months else 0.0,
        fy_budget=b["_v"].sum(), ytd_actual=ytd_actual,
        months_elapsed=months, higher_is_better=higher)))


def _signed_total(frame, line, structure):
    """The variance of a computed line on `frame`, following its own signs."""
    def value(label):
        node = structure[label]
        if node.kind == "category":
            rows = frame[frame["category"] == node.category]
            return rows["actual"].sum() - rows["budget"].sum()
        return sum((1 if s == "+" else -1) * value(r) for s, r in node.components)
    return value(line.label)


def _signed_base(frame, line, structure):
    """The same, on budget alone, as the base a percentage is taken against."""
    def value(label):
        node = structure[label]
        if node.kind == "category":
            return frame[frame["category"] == node.category]["budget"].sum()
        return sum((1 if s == "+" else -1) * value(r) for s, r in node.components)
    return value(line.label)


def _pnl_findings(txns, bud, agg, ytd_agg, perno, months):
    """Analysis for the P&L's material category lines."""
    from ..analysis import findings

    # A computed line is where the flag usually lands - EBIT and net income are
    # the two the reader stops on - and skipping them left the P&L with no
    # analysis at all. Its drivers are the lines beneath it, signed: that is the
    # level at which "what moved EBIT" has an answer.
    def line_variances(frame):
        out = {}
        for node in PNL_STRUCTURE:
            if node.kind == "category":
                rows = frame[frame["category"] == node.category]
                out[node.label] = rows["actual"].sum() - rows["budget"].sum()
            else:
                out[node.label] = sum((1 if sign == "+" else -1) * out[ref]
                                      for sign, ref in node.components)
        return out

    line_var = line_variances(agg)
    ytd_line_var = line_variances(ytd_agg)

    out = []
    for line in PNL_STRUCTURE:
        higher = line.favourable == FAV_HIGHER
        if line.kind != "category":
            _computed_block(out, line, line_var, ytd_line_var, agg, ytd_agg,
                            txns, bud, perno, months, higher)
            continue
        cat = line.category
        mrows = agg[agg["category"] == cat]
        yrows = ytd_agg[ytd_agg["category"] == cat]
        m_var = mrows["actual"].sum() - mrows["budget"].sum()
        y_var = yrows["actual"].sum() - yrows["budget"].sum()
        flag = _flag_of(m_var, _pct_or_none(m_var, mrows["budget"].sum()),
                        y_var, _pct_or_none(y_var, yrows["budget"].sum()))
        if not flag:
            continue
        drivers = yrows.copy()
        drivers["var_bud"] = drivers["actual"] - drivers["budget"]
        hist = _history(txns, bud, txns["category"] == cat,
                        bud["category"] == cat, perno)
        fy = bud[bud["category"] == cat]["budget_eur"].sum()
        ytd_actual = yrows["actual"].sum()
        out.append((line.label, flag, findings(
            drivers=drivers, total_var=y_var, history=hist,
            run_rate=ytd_actual / months * 12 if months else 0.0,
            fy_budget=fy, ytd_actual=ytd_actual, months_elapsed=months,
            higher_is_better=higher)))
    return out


def _expense_findings(txns, bud, perno, months):
    """Analysis for the expense groups that carry a flag."""
    from ..analysis import findings

    out = []
    for group_name, types in EXPENSE_GROUPS:
        tm, bm = txns["expense_type"].isin(types), bud["expense_type"].isin(types)
        month_t = txns[tm & (txns["period_no"] == perno)]
        month_b = bud[bm & (bud["period_no"] == perno)]
        ytd_t = txns[tm & (txns["period_no"] <= perno)]
        ytd_b = bud[bm & (bud["period_no"] <= perno)]
        m_var = month_t["amount_eur"].sum() - month_b["budget_eur"].sum()
        y_var = ytd_t["amount_eur"].sum() - ytd_b["budget_eur"].sum()
        flag = _flag_of(m_var, _pct_or_none(m_var, month_b["budget_eur"].sum()),
                        y_var, _pct_or_none(y_var, ytd_b["budget_eur"].sum()))
        if not flag:
            continue
        drivers = (ytd_t.groupby("expense_type", as_index=False)["amount_eur"]
                   .sum().rename(columns={"amount_eur": "actual"}))
        plan = (ytd_b.groupby("expense_type", as_index=False)["budget_eur"]
                .sum().rename(columns={"budget_eur": "budget"}))
        drivers = drivers.merge(plan, on="expense_type", how="outer").fillna(0.0)
        drivers["var_bud"] = drivers["actual"] - drivers["budget"]
        ytd_actual = ytd_t["amount_eur"].sum()
        out.append((group_name, flag, findings(
            drivers=drivers, total_var=y_var,
            history=_history(txns, bud, tm, bm, perno),
            run_rate=ytd_actual / months * 12 if months else 0.0,
            fy_budget=bud[bm]["budget_eur"].sum(), ytd_actual=ytd_actual,
            months_elapsed=months, higher_is_better=False,
            name_col="expense_type")))
    return out



def _dimension_findings(txns, bud, perno, months, *, dim, values, child,
                        higher=False, spend_only=True, net=False):
    """Analysis for every flagged member of one dimension.

    `child` is the level underneath, because that is what the concentration
    line names: a department is explained by its cost centres, an entity by its
    departments. Without a level below, there is nothing to attribute to.
    """
    from ..analysis import findings

    t = txns[txns["category"] != "Revenue"] if spend_only else txns
    b = bud[bud["category"] != "Revenue"] if spend_only else bud
    sign_t = sign_b = 1.0
    if net:
        sign_t = t["category"].eq("Revenue").map({True: 1.0, False: -1.0})
        sign_b = b["category"].eq("Revenue").map({True: 1.0, False: -1.0})
    t = t.assign(_v=t["amount_eur"] * sign_t)
    b = b.assign(_v=b["budget_eur"] * sign_b)

    out = []
    for value in values:
        tm, bm = t[dim] == value, b[dim] == value
        m_a = t.loc[tm & (t["period_no"] == perno), "_v"].sum()
        m_b = b.loc[bm & (b["period_no"] == perno), "_v"].sum()
        y_a = t.loc[tm & (t["period_no"] <= perno), "_v"].sum()
        y_b = b.loc[bm & (b["period_no"] <= perno), "_v"].sum()
        flag = _flag_of(m_a - m_b, _pct_or_none(m_a - m_b, m_b),
                        y_a - y_b, _pct_or_none(y_a - y_b, y_b))
        if not flag:
            continue
        drv = (t[tm & (t["period_no"] <= perno)].groupby(child, as_index=False)["_v"]
               .sum().rename(columns={"_v": "actual"}))
        pl = (b[bm & (b["period_no"] <= perno)].groupby(child, as_index=False)["_v"]
              .sum().rename(columns={"_v": "budget"}))
        drv = drv.merge(pl, on=child, how="outer").fillna(0.0)
        drv["var_bud"] = drv["actual"] - drv["budget"]
        hist = (t[tm & (t["period_no"] <= perno)]
                .groupby(["period", "period_no"], as_index=False)["_v"].sum()
                .rename(columns={"_v": "actual"})
                .merge(b[bm & (b["period_no"] <= perno)]
                       .groupby(["period", "period_no"], as_index=False)["_v"].sum()
                       .rename(columns={"_v": "budget"}),
                       on=["period", "period_no"], how="outer").fillna(0.0))
        out.append((value, flag, findings(
            drivers=drv, total_var=y_a - y_b, history=hist,
            run_rate=y_a / months * 12 if months else 0.0,
            fy_budget=b.loc[bm, "_v"].sum(), ytd_actual=y_a,
            months_elapsed=months, higher_is_better=higher, name_col=child)))
    return out


def _driver_findings(txns, bud, perno, months):
    """One block for the account-level sheet.

    Every other sheet's concentration line points into this one, so a block per
    account would restate them. What is not on any other sheet is how much of
    the total movement the flagged accounts carry between them.
    """
    from ..analysis import Finding

    keys = ["account_code", "account_name", "category"]
    m_a = (txns[txns["period_no"] == perno].groupby(keys, as_index=False)["amount_eur"]
           .sum().rename(columns={"amount_eur": "actual"}))
    m_b = (bud[bud["period_no"] == perno].groupby(keys, as_index=False)["budget_eur"]
           .sum().rename(columns={"budget_eur": "budget"}))
    frame = m_a.merge(m_b, on=keys, how="outer").fillna(0.0)
    frame["var_bud"] = frame["actual"] - frame["budget"]
    frame["pct"] = [_pct_or_none(v, b) for v, b in zip(frame["var_bud"], frame["budget"])]
    y_a = (txns[txns["period_no"] <= perno].groupby(keys, as_index=False)["amount_eur"]
           .sum().rename(columns={"amount_eur": "yactual"}))
    y_b = (bud[bud["period_no"] <= perno].groupby(keys, as_index=False)["budget_eur"]
           .sum().rename(columns={"budget_eur": "ybudget"}))
    frame = frame.merge(y_a, on=keys, how="left").merge(y_b, on=keys, how="left").fillna(0.0)
    frame["yvar"] = frame["yactual"] - frame["ybudget"]
    frame["ypct"] = [_pct_or_none(v, b) for v, b in zip(frame["yvar"], frame["ybudget"])]
    frame["flag"] = [_flag_of(a, b, c, d) for a, b, c, d in
                     zip(frame["var_bud"], frame["pct"], frame["yvar"], frame["ypct"])]

    flagged = frame[frame["flag"] != ""]
    if flagged.empty:
        return []
    gross = frame["var_bud"].abs().sum()
    share = flagged["var_bud"].abs().sum() / gross if gross else 0.0
    both = int((flagged["flag"] == "BOTH").sum())
    items = [Finding(
        "Coverage",
        f"{len(flagged)} of {len(frame)} accounts carry a flag, and they hold "
        f"{_pct(share)} of the gross movement on the sheet - the rest is noise "
        "around the plan.")]
    if both:
        items.append(Finding(
            "Persistence",
            f"{both} of them are adverse on the month and cumulatively, which "
            "is where a re-plan conversation starts rather than a variance "
            "explanation."))
    unfl = flagged[flagged["var_bud"] > 0] if True else flagged
    if len(unfl):
        top = unfl.reindex(unfl["var_bud"].abs().sort_values(ascending=False).index).head(3)
        items.append(Finding(
            "Largest",
            "The largest movements are "
            + _join_names([f"{r['account_name'].lower()} ({_mag(r['var_bud'])})"
                           for _, r in top.iterrows()]) + "."))
    return [("Account-level movers", "", items)]


# ===========================================================================
# P&L Report
# ===========================================================================
def _write_pnl(ws, gl, glf, gll, bud, budf, budl, period, comments, report,
               months, analysis_blocks=None) -> None:
    """The management P&L, and the sheet that owns the pack's three levers.

    Month against budget answers what happened; year to date answers whether it
    is a pattern; the run rate answers where the year lands if it continues. A
    reader asks those three questions in that order, so they sit in that order,
    and every other sheet in the pack repeats the same columns for its own cut
    of the same ledger.
    """
    hide_grid(ws)
    L = Layout(1)
    com = get_column_letter(L.ncols + 1)
    title_band(ws, "Management P&L · Variance Report (consolidated)",
               meta_line(period), com)
    gR = lambda col: f"'{gl}'!${col}${glf}:${col}${gll}"
    bR = lambda col: f"'{bud}'!${col}${budf}:${col}${budl}"
    perno = int(period[:4]) * 100 + int(period[5:7])
    mflt = f',{gR(GL_PERIOD)},"{period}"'
    mfltb = f',{bR(BUD_PERIOD)},"{period}"'
    yflt = f',{gR(GL_PERNO)},"<={perno}"'
    yfltb = f',{bR(BUD_PERNO)},"<={perno}"'

    # ---- the three levers every sheet points at ----
    lever(ws, "A9:B9", "Materiality floor (\u20ac)", LEVER_EUR,
          DEFAULT_ABS_THRESHOLD, CUR_EUR)
    lever(ws, "E9:F9", "Materiality floor (%)", LEVER_PCT,
          DEFAULT_PCT_THRESHOLD, PCT)
    # Counted from the ledger rather than read off the reporting month: an
    # extract that starts in March still reports June as month six, and dividing
    # a four-month year to date by six understates every run rate in the pack.
    lever(ws, "I9:J9", "Months elapsed", LEVER_MONTHS, months, "0")
    ws.merge_cells(f"M9:{com}9")
    ws["M9"] = ("F/U judges the month variance; Flag names which timeframe "
                "clears both floors.")
    ws["M9"].font = F_NOTE; ws["M9"].alignment = LEFT
    # The KPI cards now run to row 8, so the lever row sits directly beneath
    # them. A little more height is what keeps it from reading as a fourth line
    # of the card.
    ws.row_dimensions[9].height = 20

    headers(ws, 10, L.headers("") + ["Commentary"], center_from=2,
            center_to=L.ncols)

    subtotal_labels = {"Gross profit", "Operating income (EBIT)", "Net income"}
    label_row, first, r, cat_idx = {}, 11, 11, 0
    rows_higher, rows_lower = [], []
    for line in PNL_STRUCTURE:
        c = L.row(r)
        is_sub = line.label in subtotal_labels
        higher = line.favourable == FAV_HIGHER
        ws[f"A{r}"] = line.label
        if line.kind == "category":
            ws[c["act"]] = f'=SUMIFS({gR(GL_EUR)},{gR(GL_CAT)},"{line.category}"{mflt})'
            ws[c["bud"]] = f'=SUMIFS({bR(BUD_BUD)},{bR(BUD_CAT)},"{line.category}"{mfltb})'
            ws[c["yact"]] = f'=SUMIFS({gR(GL_EUR)},{gR(GL_CAT)},"{line.category}"{yflt})'
            ws[c["ybud"]] = f'=SUMIFS({bR(BUD_BUD)},{bR(BUD_CAT)},"{line.category}"{yfltb})'
            # No period criterion: the full-year plan is the whole budget sheet.
            ws[c["fybud"]] = f'=SUMIF({bR(BUD_CAT)},"{line.category}",{bR(BUD_BUD)})'
        else:
            parts = {key: [] for key in ("act", "bud", "yact", "ybud", "fybud")}
            for sign, ref in line.components:
                rr = label_row[ref]
                for key in parts:
                    parts[key].append(f"{sign}{getattr(L, key)}{rr}")
            for key, terms in parts.items():
                ws[c[key]] = "=" + "".join(terms)
        _tail(ws, L, r, higher=higher, bold=is_sub,
              lev_e=LEV_E, lev_p=LEV_P, months=LEV_M)

        base_fill = FILL_IVORY if is_sub else band_fill(cat_idx)
        if not is_sub:
            cat_idx += 1
        for col in L.span() + [com]:
            ws[f"{col}{r}"].fill = base_fill
        ws[f"A{r}"].font = F_SUB if is_sub else F_BODY
        ws[f"A{r}"].alignment = LEFT
        comment = comments.get(line.label, "")
        kc = ws[f"{com}{r}"]; kc.value = comment; kc.alignment = WRAP
        kc.font = F_SUB if is_sub else F_BODY
        if is_sub:
            for col in L.span() + [com]:
                ws[f"{col}{r}"].border = SUBTOTAL_TOP
        (rows_higher if higher else rows_lower).append(r)
        ws.row_dimensions[r].height = wrapped_height(comment, COMMENT_WIDTH)
        label_row[line.label] = r; r += 1
    last = r - 1
    _report_cf(ws, L, first, last, rows_higher, rows_lower)

    # ---- KPI cards ----
    # The headline is the year to date: a single month is noisy, and cumulative
    # performance against the annual plan is the figure a management P&L cover
    # is read for. The month sits under it in the same card, so the reader gets
    # the trend and the latest point without choosing between them.
    #
    # Both lines are laid out as cells with number formats rather than one cell
    # built with TEXT(). TEXT() takes its format code in the user's own
    # language, so "#,##0" is invalid on a Hungarian or German Excel and the
    # whole card rendered #VALUE! there.
    cards = [("REVENUE", "Revenue"), ("OPERATING INCOME (EBIT)", "Operating income (EBIT)"),
             ("NET INCOME", "Net income")]
    pct_cells = []
    for (title, key), (c1, c2) in zip(cards, [(1, 3), (5, 7), (9, 11)]):
        lr = label_row[key]
        Lc = get_column_letter(c1); Mid = get_column_letter(c2 - 1); Rt = get_column_letter(c2)
        for rr in (5, 6, 7, 8):
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).fill = FILL_IVORY
        for rr in (5, 6, 7):
            ws.merge_cells(f"{Lc}{rr}:{Rt}{rr}")
        ws.merge_cells(f"{Lc}8:{Mid}8")

        ws[f"{Lc}5"] = f"{title} · YTD"
        ws[f"{Lc}5"].font = F_KPI_LABEL; ws[f"{Lc}5"].alignment = LEFT
        ws[f"{Lc}6"] = f"={L.yact}{lr}"
        ws[f"{Lc}6"].font = F_KPI_VALUE; ws[f"{Lc}6"].number_format = CUR_EUR
        ws[f"{Lc}6"].alignment = LEFT
        ws[f"{Lc}7"] = f"={L.ypct}{lr}"
        ws[f"{Lc}7"].number_format = KPI_DELTA
        ws[f"{Lc}7"].font = F_KPI_DELTA; ws[f"{Lc}7"].alignment = LEFT
        # The month, muted, as the second reading rather than the headline.
        ws[f"{Lc}8"] = f"={L.act}{lr}"
        ws[f"{Lc}8"].number_format = KPI_MONTH
        ws[f"{Lc}8"].font = F_META; ws[f"{Lc}8"].alignment = LEFT
        ws[f"{Rt}8"] = f"={L.pct}{lr}"
        ws[f"{Rt}8"].number_format = KPI_MONTH_PCT
        ws[f"{Rt}8"].font = F_META; ws[f"{Rt}8"].alignment = LEFT
        outline(ws, 5, c1, 8, c2, GOLD_SIDE)
        pct_cells.append(f"{Lc}7")
    # Coloured by rule, not by a colour baked in at build time: the workbook is
    # live, so an edited input used to move the number and leave the old green
    # or red behind it. All three cards are profit or revenue lines, so higher
    # is better and the sign is enough.
    signed_variance_cf(ws, " ".join(pct_cells), True)
    ws.row_dimensions[5].height = 18; ws.row_dimensions[6].height = 26
    ws.row_dimensions[7].height = 14; ws.row_dimensions[8].height = 14

    _write_analysis(ws, L, last + 2, analysis_blocks or [], COMMENT_WIDTH)
    widths(ws, L.widths(26) + [COMMENT_WIDTH])
    quiet_indicators(ws, 4, last)
    ws.freeze_panes = L.frozen(first)


# ===========================================================================
# Expense Report (natural view: expense type)
# ===========================================================================
def _write_expense_report(ws, gl, glf, gll, bud, budf, budl, period,
                          comments=None, analysis_blocks=None) -> None:
    hide_grid(ws)
    L = Layout(1)
    com = get_column_letter(L.ncols + 1)
    comments = comments or {}
    # The meta line is built centrally, so the fiscal year follows the period
    # rather than a constant: hardcoded, this header still read "FY 2025" for a
    # 2026 close, and it phrased the same three horizons differently from the
    # sheet next to it.
    title_band(ws, "Expense Report · by expense type", meta_line(period), com)
    gR = lambda col: f"'{gl}'!${col}${glf}:${col}${gll}"
    bR = lambda col: f"'{bud}'!${col}${budf}:${col}${budl}"
    perno = int(period[:4]) * 100 + int(period[5:7])

    headers(ws, 5, L.headers("Expense Type") + ["Commentary"], center_from=2,
            center_to=L.ncols)
    _lever_echo(ws, L)

    first, r, i = 6, 6, 0
    group_rows, all_rows = [], []
    for group_name, types in EXPENSE_GROUPS:
        # A group with a single expense type needs no separate subtotal row.
        single = len(types) == 1
        type_rows = []
        for et in types:
            c = L.row(r)
            band = FILL_IVORY if single else band_fill(i)
            lab = ws.cell(row=r, column=1, value=et)
            lab.font = F_SUB if single else F_BODY
            lab.alignment = LEFT if single else indent(1)
            ws[c["act"]] = (f'=SUMIFS({gR(GL_EUR)},{gR(GL_EXP)},"{et}",'
                            f'{gR(GL_PERIOD)},"{period}")')
            ws[c["bud"]] = (f'=SUMIFS({bR(BUD_BUD)},{bR(BUD_EXP)},"{et}",'
                            f'{bR(BUD_PERIOD)},"{period}")')
            ws[c["yact"]] = (f'=SUMIFS({gR(GL_EUR)},{gR(GL_EXP)},"{et}",'
                             f'{gR(GL_PERNO)},"<={perno}")')
            ws[c["ybud"]] = (f'=SUMIFS({bR(BUD_BUD)},{bR(BUD_EXP)},"{et}",'
                             f'{bR(BUD_PERNO)},"<={perno}")')
            ws[c["fybud"]] = f'=SUMIF({bR(BUD_EXP)},"{et}",{bR(BUD_BUD)})'
            _tail(ws, L, r, higher=False, bold=single)
            for col in L.span() + [com]:
                cell = ws[f"{col}{r}"]; cell.fill = band
                if single:
                    cell.border = SUBTOTAL_TOP
            # A group holding one type is drawn as the group row, so it is a
            # roll-up on the sheet and gets the group's comment. Only the
            # "driven by" clause is dropped, because it would name this row.
            note_text = comments.get(group_name, "") if single else ""
            kc = ws[f"{com}{r}"]; kc.value = note_text
            kc.font = F_SUB if single else F_BODY; kc.alignment = WRAP
            ws.row_dimensions[r].height = max(
                21 if single else 18, wrapped_height(note_text, COMMENT_WIDTH))
            type_rows.append(r); all_rows.append(r)
            i += 1; r += 1

        if single:
            group_rows.append(type_rows[0])
            i = 0
            continue

        ws.cell(row=r, column=1, value=group_name).font = F_SUB
        ws.cell(row=r, column=1).alignment = LEFT
        _sum_tail(ws, L, r, type_rows, higher=False)
        for col in L.span() + [com]:
            cell = ws[f"{col}{r}"]; cell.fill = FILL_IVORY
            cell.border = SUBTOTAL_TOP
        # Only the groups carry a comment: a single expense type has nothing
        # underneath it to name, so the comment could only repeat its own row.
        note_text = comments.get(group_name, "")
        kc = ws[f"{com}{r}"]; kc.value = note_text
        kc.font = F_SUB; kc.alignment = WRAP
        ws.row_dimensions[r].height = max(21, wrapped_height(note_text, COMMENT_WIDTH))
        group_rows.append(r); all_rows.append(r)
        i = 0
        r += 1
    last = r - 1

    ws.cell(row=r, column=1, value="Total expenses").font = F_SUB
    ws.cell(row=r, column=1).alignment = LEFT
    _sum_tail(ws, L, r, group_rows, higher=False)
    for col in L.span() + [com]:
        ws[f"{col}{r}"].fill = FILL_IVORY
        ws[f"{col}{r}"].border = TOTAL_TOP
    total_text = comments.get(TOTAL_KEY, "")
    kc = ws[f"{com}{r}"]; kc.value = total_text
    kc.font = F_SUB; kc.alignment = WRAP
    ws.row_dimensions[r].height = max(24, wrapped_height(total_text, COMMENT_WIDTH))
    all_rows.append(r)

    _report_cf(ws, L, first, r, [], all_rows)
    ws.cell(row=r + 2, column=1,
            value="Expense types are grouped as a management report reads them: cost of "
                  "sales, then operating costs with personnel first, then non-cash and "
                  "financing items. Bold rows are group subtotals. Every line is a cost, "
                  "so an overspend is unfavourable whichever timeframe it shows up in."
            ).font = F_NOTE

    _write_analysis(ws, L, r + 4, analysis_blocks or [], COMMENT_WIDTH)
    widths(ws, L.widths(28) + [COMMENT_WIDTH])
    fit_text_columns(ws, ["A"], first, r)
    quiet_indicators(ws, 4, last + 2)
    ws.freeze_panes = L.frozen(first)


# ===========================================================================
# By Entity
# ===========================================================================
def _write_by_entity(ws, ent_var, gl, glf, gll, bud, budf, budl, period,
                     comments=None, analysis_blocks=None) -> None:
    """Net income per legal entity, on the same columns as every other sheet.

    The measure is net income - revenue less spend - because that is the one
    figure that consolidates to the group P&L. The revenue and spend halves are
    on the P&L, where they are read against the rest of the structure.
    """
    hide_grid(ws)
    L = Layout(1)
    com = get_column_letter(L.ncols + 1)
    comments = comments or {}
    title_band(ws, "By Entity · net income by legal entity", meta_line(period), com)
    headers(ws, 5, L.headers("Entity") + ["Commentary"], center_from=2,
            center_to=L.ncols)
    _lever_echo(ws, L)
    gR = lambda col: f"'{gl}'!${col}${glf}:${col}${gll}"
    bR = lambda col: f"'{bud}'!${col}${budf}:${col}${budl}"
    perno = int(period[:4]) * 100 + int(period[5:7])
    mflt = f',{gR(GL_PERIOD)},"{period}"'
    mfltb = f',{bR(BUD_PERIOD)},"{period}"'
    yflt = f',{gR(GL_PERNO)},"<={perno}"'
    yfltb = f',{bR(BUD_PERNO)},"<={perno}"'

    def net_gl(ent, flt):
        return (f'=SUMIFS({gR(GL_EUR)},{gR(GL_ENT)},"{ent}",{gR(GL_CAT)},"Revenue"{flt})'
                f'-SUMIFS({gR(GL_EUR)},{gR(GL_ENT)},"{ent}",{gR(GL_CAT)},"<>Revenue"{flt})')

    def net_bud(ent, flt):
        return (f'=SUMIFS({bR(BUD_BUD)},{bR(BUD_ENT)},"{ent}",{bR(BUD_CAT)},"Revenue"{flt})'
                f'-SUMIFS({bR(BUD_BUD)},{bR(BUD_ENT)},"{ent}",{bR(BUD_CAT)},"<>Revenue"{flt})')

    first, r, rows = 6, 6, []
    for i, (_, row) in enumerate(ent_var.iterrows()):
        ent = row["entity"]; band = band_fill(i)
        c = L.row(r)
        ws.cell(row=r, column=1, value=ent).font = F_BODY
        ws.cell(row=r, column=1).alignment = LEFT
        ws[c["act"]] = net_gl(ent, mflt)
        ws[c["bud"]] = net_bud(ent, mfltb)
        ws[c["yact"]] = net_gl(ent, yflt)
        ws[c["ybud"]] = net_bud(ent, yfltb)
        ws[c["fybud"]] = net_bud(ent, "")
        _tail(ws, L, r, higher=True)
        for col in L.span() + [com]:
            ws[f"{col}{r}"].fill = band
        note_text = comments.get(ent, "")
        kc = ws[f"{com}{r}"]; kc.value = note_text
        kc.font = F_BODY; kc.alignment = WRAP
        ws.row_dimensions[r].height = max(22, wrapped_height(note_text, COMMENT_WIDTH))
        rows.append(r); r += 1
    last = r - 1

    ws.cell(row=r, column=1, value="Consolidated").font = F_SUB
    ws.cell(row=r, column=1).alignment = LEFT
    _sum_tail(ws, L, r, rows, higher=True)
    for col in L.span() + [com]:
        ws[f"{col}{r}"].fill = FILL_IVORY
        ws[f"{col}{r}"].border = TOTAL_TOP
    total_text = comments.get(TOTAL_KEY, "")
    kc = ws[f"{com}{r}"]; kc.value = total_text
    kc.font = F_SUB; kc.alignment = WRAP
    ws.row_dimensions[r].height = max(24, wrapped_height(total_text, COMMENT_WIDTH))
    rows.append(r)

    # Ranges run to `r`, the total row, not `last`: a total carries the same
    # verdict as the lines above it and was being left uncoloured.
    _report_cf(ws, L, first, r, rows, [])
    ws.cell(row=r + 2, column=1,
            value="Each entity's net income is its revenue less its spend, which is why "
                  "the entities consolidate to the group P&L. The revenue and spend "
                  "halves are on the P&L Report.").font = F_NOTE
    _write_analysis(ws, L, r + 4, analysis_blocks or [], COMMENT_WIDTH)
    widths(ws, L.widths(18) + [COMMENT_WIDTH])
    fit_text_columns(ws, ["A"], first, r)
    quiet_indicators(ws, 4, last + 2)
    ws.freeze_panes = L.frozen(first)


# ===========================================================================
# Cost Centres
# ===========================================================================
def _write_cost_centres(ws, dept_var, gl, glf, gll, bud, budf, budl, period,
                        comments=None, analysis_blocks=None) -> None:
    """Departmental spend variance with each department's cost centres nested."""
    hide_grid(ws)
    L = Layout(1)
    com = get_column_letter(L.ncols + 1)
    comments = comments or {}
    title_band(ws, "Departments & Cost Centres · spend variance", meta_line(period), com)
    headers(ws, 5, L.headers("Department / Cost Centre") + ["Commentary"],
            center_from=2, center_to=L.ncols)
    _lever_echo(ws, L)

    gR = lambda col: f"'{gl}'!${col}${glf}:${col}${gll}"
    bR = lambda col: f"'{bud}'!${col}${budf}:${col}${budl}"
    perno = int(period[:4]) * 100 + int(period[5:7])
    mflt = f',{gR(GL_PERIOD)},"{period}"'
    mfltb = f',{bR(BUD_PERIOD)},"{period}"'
    yflt = f',{gR(GL_PERNO)},"<={perno}"'
    yfltb = f',{bR(BUD_PERNO)},"<={perno}"'

    order = dept_var.set_index("department")["budget"].to_dict()
    depts = sorted(SPEND_DEPARTMENTS, key=lambda d: order.get(d, 0), reverse=True)

    def spend_gl(col, key, flt):
        return f'=SUMIFS({gR(GL_EUR)},{gR(col)},"{key}",{gR(GL_CAT)},"<>Revenue"{flt})'

    def spend_bud(col, key, flt):
        return f'=SUMIFS({bR(BUD_BUD)},{bR(col)},"{key}",{bR(BUD_CAT)},"<>Revenue"{flt})'

    first, r, dept_rows, all_rows = 6, 6, [], []
    for dept in depts:
        c = L.row(r)
        ws.cell(row=r, column=1, value=dept).font = F_SUB
        ws.cell(row=r, column=1).alignment = LEFT
        ws[c["act"]] = spend_gl(GL_DEPT, dept, mflt)
        ws[c["bud"]] = spend_bud(BUD_DEPT, dept, mfltb)
        ws[c["yact"]] = spend_gl(GL_DEPT, dept, yflt)
        ws[c["ybud"]] = spend_bud(BUD_DEPT, dept, yfltb)
        ws[c["fybud"]] = spend_bud(BUD_DEPT, dept, "")
        _tail(ws, L, r, higher=False, bold=True)
        for col in L.span() + [com]:
            ws[f"{col}{r}"].fill = FILL_IVORY
            ws[f"{col}{r}"].border = SUBTOTAL_TOP
        # The department is the roll-up, so it is the row that can name which of
        # its cost centres moved the total. The cost centres below it are leaves
        # and get nothing: there the comment could only restate the row.
        note_text = comments.get(dept, "")
        kc = ws[f"{com}{r}"]; kc.value = note_text
        kc.font = F_SUB; kc.alignment = WRAP
        ws.row_dimensions[r].height = max(22, wrapped_height(note_text, COMMENT_WIDTH))
        dept_rows.append(r); all_rows.append(r); r += 1

        for i, (cc_code, cc_name) in enumerate(DEPARTMENT_COST_CENTRES[dept]):
            c = L.row(r)
            label = ws.cell(row=r, column=1, value=f"{cc_code}   {cc_name}")
            label.font = F_BODY
            label.alignment = indent(2)
            ws[c["act"]] = spend_gl(GL_CC, cc_code, mflt)
            ws[c["bud"]] = spend_bud(BUD_CC, cc_code, mfltb)
            ws[c["yact"]] = spend_gl(GL_CC, cc_code, yflt)
            ws[c["ybud"]] = spend_bud(BUD_CC, cc_code, yfltb)
            ws[c["fybud"]] = spend_bud(BUD_CC, cc_code, "")
            _tail(ws, L, r, higher=False)
            band = band_fill(i)
            for col in L.span() + [com]:
                ws[f"{col}{r}"].fill = band
            ws.row_dimensions[r].height = 18
            all_rows.append(r); r += 1
    last = r - 1

    # --- grand total: sum of the department rows only (no double counting) ---
    ws.cell(row=r, column=1, value="Total spend").font = F_SUB
    ws.cell(row=r, column=1).alignment = LEFT
    _sum_tail(ws, L, r, dept_rows, higher=False)
    for col in L.span() + [com]:
        ws[f"{col}{r}"].fill = FILL_IVORY
        ws[f"{col}{r}"].border = TOTAL_TOP
    total_text = comments.get(TOTAL_KEY, "")
    kc = ws[f"{com}{r}"]; kc.value = total_text
    kc.font = F_SUB; kc.alignment = WRAP
    ws.row_dimensions[r].height = max(24, wrapped_height(total_text, COMMENT_WIDTH))
    all_rows.append(r)

    _report_cf(ws, L, first, r, [], all_rows)
    ws.cell(row=r + 2, column=1,
            value="Departments are bold roll-ups; cost centres are indented beneath their "
                  "department, shown as code plus description. Totals sum the departments "
                  "only. Revenue is excluded: this is a spend view.").font = F_NOTE

    _write_analysis(ws, L, r + 5, analysis_blocks or [], COMMENT_WIDTH)
    widths(ws, L.widths(32) + [COMMENT_WIDTH])
    fit_text_columns(ws, ["A"], first, r)
    quiet_indicators(ws, 4, last + 3)
    ws.freeze_panes = L.frozen(first)


# ===========================================================================
# Drivers
# ===========================================================================
def _write_drivers(ws, agg, gl, glf, gll, bud, budf, budl, period,
                   analysis_blocks=None) -> None:
    hide_grid(ws)
    L = Layout(3)
    title_band(ws, "Variance Drivers · account level", meta_line(period), L.last_col)
    headers(ws, 5, L.headers("Account", "Name", "Category"),
            center_from=4, center_to=L.ncols)
    _lever_echo(ws, L)
    gR = lambda col: f"'{gl}'!${col}${glf}:${col}${gll}"
    bR = lambda col: f"'{bud}'!${col}${budf}:${col}${budl}"
    perno = int(period[:4]) * 100 + int(period[5:7])
    mflt = f',{gR(GL_PERIOD)},"{period}"'
    mfltb = f',{bR(BUD_PERIOD)},"{period}"'
    yflt = f',{gR(GL_PERNO)},"<={perno}"'
    yfltb = f',{bR(BUD_PERNO)},"<={perno}"'

    tmp = agg.copy(); tmp["_var"] = tmp["actual"] - tmp["budget"]
    tmp = tmp.sort_values("_var", key=lambda s: s.abs(), ascending=False).reset_index(drop=True)
    first, r = 6, 6
    rows_higher, rows_lower = [], []
    for i, (_, row) in enumerate(tmp.iterrows()):
        higher = CATEGORY_FAVOURABLE[row["category"]] == FAV_HIGHER
        code = row["account_code"]
        band = band_fill(i)
        c = L.row(r)
        ws.cell(row=r, column=1, value=code).font = F_BODY
        ws.cell(row=r, column=2, value=row["account_name"]).font = F_BODY
        ws.cell(row=r, column=3, value=row["category"]).font = F_BODY
        ws[c["act"]] = f'=SUMIFS({gR(GL_EUR)},{gR(GL_CODE)},"{code}"{mflt})'
        ws[c["bud"]] = f'=SUMIFS({bR(BUD_BUD)},{bR(BUD_CODE)},"{code}"{mfltb})'
        ws[c["yact"]] = f'=SUMIFS({gR(GL_EUR)},{gR(GL_CODE)},"{code}"{yflt})'
        ws[c["ybud"]] = f'=SUMIFS({bR(BUD_BUD)},{bR(BUD_CODE)},"{code}"{yfltb})'
        ws[c["fybud"]] = f'=SUMIF({bR(BUD_CODE)},"{code}",{bR(BUD_BUD)})'
        _tail(ws, L, r, higher=higher)
        for col in L.span():
            ws[f"{col}{r}"].fill = band
        for col in ("A", "B", "C"):
            ws[f"{col}{r}"].alignment = LEFT
        ws.row_dimensions[r].height = 18
        (rows_higher if higher else rows_lower).append(r)
        r += 1
    last = r - 1
    # No data bar on the variance column: the rows are already ordered by the
    # size of the movement, while a bar encodes the signed value, so the longest
    # bar was rarely the top row. The colour carries the direction, the F/U cell
    # carries the verdict and the flag carries the timeframe - a fourth encoding
    # on the same cell only competed with the digits behind it.
    _report_cf(ws, L, first, last, rows_higher, rows_lower)
    _write_analysis(ws, L, last + 3, analysis_blocks or [], COMMENT_WIDTH)
    widths(ws, L.widths(12, 30, 12))
    fit_text_columns(ws, ["B", "C"], first, last)
    quiet_indicators(ws, 4, last)
    ws.freeze_panes = L.frozen(first)

def demo_ytd_detail(period: str, seed: int = 42):
    """The sample company's cumulative detail, for the deck's spend slide."""
    from ..synthetic_data import generate_ytd_transactions, generate_budget_year
    perno = int(period[:4]) * 100 + int(period[5:7])
    return _ytd_detail(generate_ytd_transactions(period, seed),
                       generate_budget_year(seed), perno)


def demo_analysis_blocks(period: str, seed: int = 42) -> list:
    """The findings the deck shows, built from the same generated ledger.

    Exposed so the deck and the workbook cannot analyse the same company two
    different ways: both call this.
    """
    from ..synthetic_data import (generate_ytd_transactions, generate_budget_year,
                                  generate_month)
    perno = int(period[:4]) * 100 + int(period[5:7])
    txns = generate_ytd_transactions(period, seed)
    bud = generate_budget_year(seed)
    agg = generate_month(period, seed).drop(columns="period")
    ytd_detail = _ytd_detail(txns, bud, perno)
    ytd_agg = (ytd_detail.groupby(["account_code", "account_name", "category"],
                                  as_index=False)[["actual", "budget"]].sum())
    ytd_agg["prior_year"] = 0.0
    months = txns.loc[txns["period_no"] <= perno, "period_no"].nunique() or 1
    return (_pnl_findings(txns, bud, agg, ytd_agg, perno, months)
            + _expense_findings(txns, bud, perno, months))


# ===========================================================================
def build_demo_pack(period: str, out_path: str | Path, seed: int = 42) -> Path:
    """Build the full multi-entity showcase pack from generated data."""
    from ..synthetic_data import (generate_ytd_transactions, generate_budget_year,
                                  monthly_detail, generate_month)
    from ..engine import department_variances, entity_variances

    txns = generate_ytd_transactions(period, seed)
    bud = generate_budget_year(seed).sort_values(
        ["period_no", "entity", "department", "cost_centre", "account_code"]
    ).reset_index(drop=True)
    detail = monthly_detail(period, seed)
    agg = generate_month(period, seed).drop(columns="period")

    perno = int(period[:4]) * 100 + int(period[5:7])
    ytd_detail = _ytd_detail(txns, bud, perno)
    ytd_agg = (ytd_detail.groupby(["account_code", "account_name", "category"],
                                  as_index=False)[["actual", "budget"]].sum())
    # The engine reports a prior-year column; the year-to-date frame is built
    # for the commentary only and carries no prior year, so it is zeroed rather
    # than left missing.
    ytd_agg["prior_year"] = 0.0

    report = build_report(agg)
    ytd_report = build_report(ytd_agg)
    # The commentary covers both timeframes, so it can explain the flag beside
    # it rather than restating one of the columns.
    comments = line_comments(report, agg, ytd_report=ytd_report, ytd_gl=ytd_agg)
    spend, ytd_spend = _spend_detail(detail), _spend_detail(ytd_detail)
    dept_comments = rollup_comments(spend, "department", "cost_centre",
                                    child_names=CC_NAMES, ytd_detail=ytd_spend)
    ent_comments = rollup_comments(_net_detail(detail), "entity", "department",
                                   higher_is_better=True,
                                   ytd_detail=_net_detail(ytd_detail))
    exp_comments = rollup_comments(_with_group(spend), "expense_group",
                                   "expense_type", ytd_detail=_with_group(ytd_spend))
    # The grand totals are roll-ups too, over every row on their sheet.
    dept_comments["\u0000total"] = total_comment(
        spend, "department", ytd_detail=ytd_spend)
    ent_comments["\u0000total"] = total_comment(
        _net_detail(detail), "entity", higher_is_better=True,
        ytd_detail=_net_detail(ytd_detail))
    exp_comments["\u0000total"] = total_comment(
        _with_group(spend), "expense_group", ytd_detail=_with_group(ytd_spend))
    months = txns.loc[txns["period_no"] <= perno, "period_no"].nunique() or 1
    pnl_analysis = _pnl_findings(txns, bud, agg, ytd_agg, perno, months)
    exp_analysis = _expense_findings(txns, bud, perno, months)
    ent_analysis = _dimension_findings(
        txns, bud, perno, months, dim="entity",
        values=sorted(txns["entity"].unique()), child="department",
        higher=True, spend_only=False, net=True)
    dept_analysis = _dimension_findings(
        txns, bud, perno, months, dim="department",
        values=SPEND_DEPARTMENTS, child="cost_centre")
    drv_analysis = _driver_findings(txns, bud, perno, months)
    dept_var = department_variances(detail)
    ent_var = entity_variances(detail)

    wb = Workbook()
    ws_gl = wb.active; ws_gl.title = "GL Transactions"
    glf, gll = _write_gl_transactions(ws_gl, txns, period)
    ws_bud = wb.create_sheet("Budget"); budf, budl = _write_budget(ws_bud, bud, period)

    GL, BUD = "GL Transactions", "Budget"
    _write_pnl(wb.create_sheet("P&L Report"), GL, glf, gll, BUD, budf, budl,
               period, comments, report, months, pnl_analysis)
    _write_expense_report(wb.create_sheet("Expense Report"), GL, glf, gll, BUD,
                          budf, budl, period, exp_comments, exp_analysis)
    _write_by_entity(wb.create_sheet("By Entity"), ent_var, GL, glf, gll, BUD,
                     budf, budl, period, ent_comments, ent_analysis)
    _write_cost_centres(wb.create_sheet("Departments & CCs"), dept_var, GL, glf,
                        gll, BUD, budf, budl, period, dept_comments, dept_analysis)
    _write_drivers(wb.create_sheet("Drivers"), agg, GL, glf, gll, BUD, budf, budl,
                   period, drv_analysis)

    desired = ["P&L Report", "Expense Report", "By Entity",
               "Departments & CCs", "Drivers", "Budget", "GL Transactions"]
    for i, name in enumerate(desired):
        wb.move_sheet(name, -wb.sheetnames.index(name) + i)
    wb.active = wb.sheetnames.index("P&L Report")

    out_path = Path(out_path); out_path.parent.mkdir(parents=True, exist_ok=True)
    quiet = collect_quiet_ranges(wb)
    wb.save(out_path)
    suppress_error_indicators(out_path, quiet)
    return out_path


if __name__ == "__main__":  # pragma: no cover - manual run
    root = Path(__file__).resolve().parents[3]
    out = build_demo_pack("2025-06", root / "output" / "flux_demo_pack.xlsx")
    print(f"Written: {out}")
