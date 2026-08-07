"""
Excel pack for ingested client data.

The sheets adapt to whichever dimensions the uploaded file actually carries:

    always            GL Input, P&L Report, Drivers
    + expense_type    Expense Report      (natural view)
    + department      Departments & CCs   (functional view, hierarchy taken
                                           from the data, not from a fixed
                                           chart of accounts)
    + entity          By Entity           (consolidation)

Every figure is a live formula over the GL Input sheet, so the pack recalculates
when an input is edited. Styling helpers are shared with excel_export.
"""

from __future__ import annotations
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

from ..coa import PNL_STRUCTURE, CATEGORY_FAVOURABLE, FAV_HIGHER, EXPENSE_GROUPS
from ..commentary import line_comments
from ..engine import build_report, has_budget as _detect_budget
from .styling import (
    FONT, GREEN_INK, RED_INK, BAR,
    F_BODY, F_SMALL, F_SUB, F_INPUT, F_LABEL, F_KPI_LABEL, F_KPI_VALUE, F_NOTE,
    F_FU, F_FLAG,
    FILL_IVORY, FILL_BAND, FILL_WHITE, FILL_LEVER,
    CUR, CUR_EUR, PCT, KPI_DELTA,
    CENTER, LEFT, RIGHT, WRAP, indent, band_fill,
    GOLD_SIDE, SUBTOTAL_TOP, TOTAL_TOP,
    hide_grid, title_band, headers, widths, outline, badge_cf, note, wrapped_height,
    variance_cf, spend_variance_cf,
    fit_text_columns,
    quiet_indicators, collect_quiet_ranges, suppress_error_indicators,
)
from .formulas import (
    LEVER_EUR, LEVER_PCT, LEV_E, LEV_P, PL_E, PL_P,
    DEFAULT_ABS_THRESHOLD, DEFAULT_PCT_THRESHOLD,
    pct_f, Variance, NO_BUDGET_NOTE,
)


OPTIONAL_DIMS = [("expense_type", "Expense type", 22),
                 ("department", "Department", 16),
                 ("cost_centre", "Cost centre", 14),
                 ("entity", "Entity", 16)]


def _present(df: pd.DataFrame, col: str) -> bool:
    """True when a dimension exists and actually carries values."""
    if col not in df.columns:
        return False
    vals = df[col].astype(str).str.strip()
    return bool((vals != "").any() and (vals != "(multiple)").any())


# ---------------------------------------------------------------------------
# GL Input
# ---------------------------------------------------------------------------
def _gl_input(ws, agg, period, dims, budgeted=True):
    hide_grid(ws)
    spec = []
    if "period" in dims:
        spec += [("period", "Period", 11), ("period_no", "Per. key", 10)]
    spec += [("account_code", "Account", 12), ("account_name", "Account name", 30),
             ("category", "Category", 12)]
    spec += [(k, label, w) for k, label, w in OPTIONAL_DIMS if k in dims]
    spec += [("actual", "Actual", 15), ("budget", "Budget", 15),
             ("prior_year", "Prior Yr Act", 15)]
    letters = {k: get_column_letter(i) for i, (k, _l, _w) in enumerate(spec, start=1)}
    money = {"actual", "budget", "prior_year"}
    ints = {"period_no"}

    title_band(ws, "GL Input · account level", f"Reporting month {period}  ·  \u20ac",
                   get_column_letter(len(spec)))
    headers(ws, 5, [l for _k, l, _w in spec],
                center_from=len(spec) - 2, center_to=len(spec))
    for i, (_k, _l, w) in enumerate(spec, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    first = 6
    for i, (_, r) in enumerate(agg.iterrows()):
        row = first + i
        band = band_fill(i)
        for c, (key, _l, _w) in enumerate(spec, start=1):
            v = r.get(key, "")
            if key in money:
                cell = ws.cell(row=row, column=c, value=float(v or 0))
                cell.number_format = CUR_EUR; cell.font = F_INPUT; cell.alignment = RIGHT
            elif key in ints:
                cell = ws.cell(row=row, column=c, value=int(v) if pd.notna(v) else None)
                cell.number_format = "0"; cell.font = F_SMALL; cell.alignment = RIGHT
            else:
                cell = ws.cell(row=row, column=c, value="" if pd.isna(v) else str(v))
                cell.font = F_BODY; cell.alignment = LEFT
            cell.fill = band
        ws.row_dimensions[row].height = 17
    last = first + len(agg) - 1
    if not budgeted:
        note(ws, last + 2, NO_BUDGET_NOTE)
    quiet_indicators(ws, 5, last)
    ws.freeze_panes = f"A{first}"
    return first, last, letters


# ---------------------------------------------------------------------------
# P&L
# ---------------------------------------------------------------------------
def _pnl(ws, gl, glf, gll, L, period, comments, report, var, perno=None, ytd=False):
    hide_grid(ws)
    meta = (f"Reporting month {period}  ·  with year to date  ·  \u20ac" if ytd
            else f"Reporting month {period}  ·  \u20ac")
    title_band(ws, "Management P&L · Variance Report", meta, "L" if ytd else "K")
    R = lambda c: f"'{gl}'!${c}${glf}:${c}${gll}"
    cat, act, bud, pri = R(L["category"]), R(L["actual"]), R(L["budget"]), R(L["prior_year"])
    # Month figures filter to the reporting period; YTD sums everything up to it.
    mflt = f',{R(L["period_no"])},{perno}' if perno and "period_no" in L else ""
    yflt = f',{R(L["period_no"])},"<={perno}"' if perno and "period_no" in L else ""

    # The two materiality thresholds are editable lever cells that every other
    # sheet points at. With no budget nothing can be flagged, so they are left
    # off rather than shown as controls that do nothing.
    if var:
        ws.merge_cells("A9:B9"); ws["A9"] = "Materiality floor (\u20ac)"
        ws["A9"].font = F_LABEL; ws["A9"].alignment = LEFT
        ws[LEVER_EUR] = DEFAULT_ABS_THRESHOLD; ws[LEVER_EUR].font = F_INPUT
        ws[LEVER_EUR].number_format = CUR_EUR; ws[LEVER_EUR].fill = FILL_LEVER
        ws[LEVER_EUR].alignment = CENTER
        ws.merge_cells("E9:F9"); ws["E9"] = "Materiality floor (%)"
        ws["E9"].font = F_LABEL; ws["E9"].alignment = LEFT
        ws[LEVER_PCT] = DEFAULT_PCT_THRESHOLD; ws[LEVER_PCT].font = F_INPUT
        ws[LEVER_PCT].number_format = PCT; ws[LEVER_PCT].fill = FILL_LEVER
        ws[LEVER_PCT].alignment = CENTER
        outline(ws, 9, 3, 9, 3, GOLD_SIDE); outline(ws, 9, 7, 9, 7, GOLD_SIDE)

    heads = (["", "Month Act", "Month Bud", "Month Var", "Var %",
              "YTD Act", "YTD Bud", "YTD Var", "Var %", "F/U", "Flag", "Commentary"]
             if ytd else
             ["", "Actual", "Budget", "Var (Bud)", "Var %", "Prior Yr Act",
              "Var (PY)", "Var %", "F/U", "Flag", "Commentary"])
    headers(ws, 10, heads, center_from=2, center_to=len(heads) - 1)
    fu_col, flag_col, com_col = (("J", "K", "L") if ytd else ("I", "J", "K"))
    # Declared before the rows are written: the commentary row height is
    # estimated from the width of the column the text has to wrap inside.
    width_list = ([26, 13, 13, 12, 9, 13, 13, 12, 9, 6, 11, 52] if ytd
                  else [26, 13, 13, 13, 9, 13, 13, 9, 6, 11, 58])
    subs = {"Gross profit", "Operating income (EBIT)", "Net income"}
    label_row, first, r, ci = {}, 11, 11, 0
    for line in PNL_STRUCTURE:
        A, B, C, D, E = f"A{r}", f"B{r}", f"C{r}", f"D{r}", f"E{r}"
        Fc, G, H = f"F{r}", f"G{r}", f"H{r}"
        I, J, K = f"{fu_col}{r}", f"{flag_col}{r}", f"{com_col}{r}"
        is_sub = line.label in subs
        ws[A] = line.label
        if line.kind == "category":
            ws[B] = f'=SUMIFS({act},{cat},"{line.category}"{mflt})'
            ws[C] = var.cell(f'=SUMIFS({bud},{cat},"{line.category}"{mflt})')
            if ytd:
                ws[Fc] = f'=SUMIFS({act},{cat},"{line.category}"{yflt})'
                ws[G] = var.cell(f'=SUMIFS({bud},{cat},"{line.category}"{yflt})')
            else:
                ws[Fc] = f'=SUMIFS({pri},{cat},"{line.category}"{mflt})'
        else:
            pb, pc, pf, pg = [], [], [], []
            for sign, ref in line.components:
                rr = label_row[ref]
                pb.append(f"{sign}B{rr}"); pc.append(f"{sign}C{rr}")
                pf.append(f"{sign}F{rr}"); pg.append(f"{sign}G{rr}")
            ws[B] = "=" + "".join(pb)
            ws[C] = var.cell("=" + "".join(pc))
            ws[Fc] = "=" + "".join(pf)
            if ytd:
                ws[G] = var.cell("=" + "".join(pg))
        ws[D] = var.cell(f"={B}-{C}"); ws[E] = var.pct(D, C)
        if ytd:
            ws[H] = var.cell(f"={Fc}-{G}")
            ws[f"I{r}"] = var.pct(H, G)
        else:
            ws[G] = f"={B}-{Fc}"; ws[H] = pct_f(G, Fc)
        ws[I] = var.fu(D, line.favourable == FAV_HIGHER)
        ws[J] = var.flag(D, E, LEV_E, LEV_P)

        fill = FILL_IVORY if is_sub else (FILL_BAND if ci % 2 else FILL_WHITE)
        if not is_sub:
            ci += 1
        span = "ABCDEFGHIJKL" if ytd else "ABCDEFGHIJK"
        for col in span:
            ws[f"{col}{r}"].fill = fill
        money_cols = ("B", "C", "D", "F", "G", "H") if ytd else ("B", "C", "D", "F", "G")
        pct_cols = ("E", "I") if ytd else ("E", "H")
        for col in money_cols:
            cc = ws[f"{col}{r}"]; cc.number_format = CUR; cc.alignment = RIGHT
            cc.font = F_SUB if is_sub else F_BODY
        for col in pct_cols:
            cc = ws[f"{col}{r}"]; cc.number_format = PCT; cc.alignment = RIGHT
            cc.font = F_SUB if is_sub else F_BODY
        ws[A].font = F_SUB if is_sub else F_BODY; ws[A].alignment = LEFT
        ws[I].alignment = CENTER; ws[I].font = F_FU
        ws[J].alignment = CENTER; ws[J].font = F_FLAG
        comment = comments.get(line.label, "")
        kc = ws[K]; kc.value = comment; kc.alignment = WRAP
        kc.font = F_SUB if is_sub else F_BODY
        if is_sub:
            for col in span:
                ws[f"{col}{r}"].border = SUBTOTAL_TOP
        ws.row_dimensions[r].height = wrapped_height(comment, width_list[-1])
        label_row[line.label] = r; r += 1
    last = r - 1
    badge_cf(ws, f"{fu_col}{first}:{fu_col}{last}", f"{flag_col}{first}:{flag_col}{last}")
    if var:
        variance_cf(ws, f"D{first}:E{last}", f"${fu_col}{first}")
        if ytd:
            variance_cf(ws, f"H{first}:I{last}", f"${fu_col}{first}")

    for (title, key), (c1, c2) in zip(
        [("REVENUE", "Revenue"), ("OPERATING INCOME (EBIT)", "Operating income (EBIT)"),
         ("NET INCOME", "Net income")],
        [(1, 3), (5, 7), (9, 12)] if ytd else [(1, 3), (5, 7), (9, 11)]):
        lr = label_row[key]; fav = report[report["line"] == key].iloc[0]["fav_unfav"] == "F"
        Lc = get_column_letter(c1); Rt = get_column_letter(c2)
        for rr in (5, 6, 7):
            ws.merge_cells(f"{Lc}{rr}:{Rt}{rr}")
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).fill = FILL_IVORY
        ws[f"{Lc}5"] = title; ws[f"{Lc}5"].font = F_KPI_LABEL
        ws[f"{Lc}5"].alignment = LEFT
        ws[f"{Lc}6"] = f"=B{lr}"; ws[f"{Lc}6"].font = F_KPI_VALUE
        ws[f"{Lc}6"].number_format = CUR_EUR
        ws[f"{Lc}6"].alignment = LEFT
        ws[f"{Lc}7"] = var.cell(f"=E{lr}")
        ws[f"{Lc}7"].number_format = KPI_DELTA
        ws[f"{Lc}7"].font = Font(name=FONT, size=9, bold=True,
                                 color=GREEN_INK if fav else RED_INK)
        ws[f"{Lc}7"].alignment = LEFT
        outline(ws, 5, c1, 7, c2, GOLD_SIDE)
    ws.row_dimensions[5].height = 18; ws.row_dimensions[6].height = 26; ws.row_dimensions[7].height = 16
    for c, w in enumerate(width_list, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    if not var:
        note(ws, last + 2, NO_BUDGET_NOTE)
    quiet_indicators(ws, 5, last + 2)
    ws.freeze_panes = f"A{first}"


# ---------------------------------------------------------------------------
# Generic one-dimension variance sheet (expense type / entity)
# ---------------------------------------------------------------------------
def _dimension_sheet(ws, values, gl, glf, gll, L, period, title, first_header,
                     dim_key, var, spend_only=False, revenue_split=False,
                     perno=None, ytd=False):
    hide_grid(ws)
    meta = (f"Reporting month {period}  ·  with year to date  ·  \u20ac" if ytd
            else f"Reporting month {period}  ·  \u20ac")
    title_band(ws, title, meta, "K" if ytd else "H")
    R = lambda c: f"'{gl}'!${c}${glf}:${c}${gll}"
    dim, cat = R(L[dim_key]), R(L["category"])
    act, bud = R(L["actual"]), R(L["budget"])
    mflt = f',{R(L["period_no"])},{perno}' if perno and "period_no" in L else ""
    yflt = f',{R(L["period_no"])},"<={perno}"' if perno and "period_no" in L else ""

    if revenue_split:
        heads = [first_header, "Revenue", "Spend", "Net income", "Budget net",
                 "Var (Bud)", "Var %", "F/U"]
    elif ytd:
        heads = [first_header, "Month Act", "Month Bud", "Month Var", "Var %",
                 "YTD Act", "YTD Bud", "YTD Var", "Var %", "F/U", "Flag"]
    else:
        heads = [first_header, "Actual", "Budget", "Var (Bud)", "Var %", "F/U", "Flag", ""]
    headers(ws, 5, heads, center_from=2, center_to=len(heads))

    first, r = 6, 6
    for i, v in enumerate(values):
        band = band_fill(i)
        ws.cell(row=r, column=1, value=v).font = F_BODY
        ws.cell(row=r, column=1).alignment = LEFT
        if revenue_split:
            B, C, D, E, Fc, G, H = (f"B{r}", f"C{r}", f"D{r}", f"E{r}", f"F{r}", f"G{r}", f"H{r}")
            ws[B] = f'=SUMIFS({act},{dim},"{v}",{cat},"Revenue"{mflt})'
            ws[C] = f'=SUMIFS({act},{dim},"{v}",{cat},"<>Revenue"{mflt})'
            ws[D] = f"={B}-{C}"
            ws[E] = var.cell(f'=SUMIFS({bud},{dim},"{v}",{cat},"Revenue"{mflt})'
                             f'-SUMIFS({bud},{dim},"{v}",{cat},"<>Revenue"{mflt})')
            ws[Fc] = var.cell(f"={D}-{E}"); ws[G] = var.pct(Fc, E); ws[H] = var.fu(Fc, True)
            for col in (B, C, D, E, Fc):
                ws[col].number_format = CUR; ws[col].font = F_BODY; ws[col].alignment = RIGHT
            ws[G].number_format = PCT; ws[G].font = F_BODY; ws[G].alignment = RIGHT
            ws[H].alignment = CENTER; ws[H].font = F_FU
            ncols = 8
        else:
            crit = f',{cat},"<>Revenue"' if spend_only else ""
            B, C, D, E = f"B{r}", f"C{r}", f"D{r}", f"E{r}"
            ws[B] = f'=SUMIFS({act},{dim},"{v}"{crit}{mflt})'
            ws[C] = var.cell(f'=SUMIFS({bud},{dim},"{v}"{crit}{mflt})')
            ws[D] = var.cell(f"={B}-{C}"); ws[E] = var.pct(D, C)
            if ytd:
                F2, G2, H2, I2 = f"F{r}", f"G{r}", f"H{r}", f"I{r}"
                ws[F2] = f'=SUMIFS({act},{dim},"{v}"{crit}{yflt})'
                ws[G2] = var.cell(f'=SUMIFS({bud},{dim},"{v}"{crit}{yflt})')
                ws[H2] = var.cell(f"={F2}-{G2}"); ws[I2] = var.pct(H2, G2)
                Fc, G = f"J{r}", f"K{r}"
                money, pcts, ncols = (B, C, D, F2, G2, H2), (E, I2), 11
            else:
                Fc, G = f"F{r}", f"G{r}"
                money, pcts, ncols = (B, C, D), (E,), 7
            ws[Fc] = var.fu(D, not spend_only)
            ws[G] = var.flag(D, E, PL_E, PL_P)
            for col in money:
                ws[col].number_format = CUR; ws[col].font = F_BODY; ws[col].alignment = RIGHT
            for col in pcts:
                ws[col].number_format = PCT; ws[col].font = F_BODY; ws[col].alignment = RIGHT
            ws[Fc].alignment = CENTER; ws[Fc].font = F_FU
            ws[G].alignment = CENTER; ws[G].font = F_FLAG
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = band
        ws.cell(row=r, column=1).alignment = LEFT
        ws.row_dimensions[r].height = 19
        r += 1
    last = r - 1

    ws.cell(row=r, column=1, value="Total").font = F_SUB
    if revenue_split:
        for col in ("B", "C"):
            ws[f"{col}{r}"] = f"=SUM({col}{first}:{col}{last})"
        ws[f"E{r}"] = var.cell(f"=SUM(E{first}:E{last})")
        ws[f"D{r}"] = f"=B{r}-C{r}"; ws[f"F{r}"] = var.cell(f"=D{r}-E{r}")
        ws[f"G{r}"] = var.pct(f"F{r}", f"E{r}"); ws[f"H{r}"] = var.fu(f"F{r}", True)
        num, pct_col, fu_col, ncols = ("B", "C", "D", "E", "F"), "G", "H", 8
    elif ytd:
        for col in ("B", "F"):
            ws[f"{col}{r}"] = f"=SUM({col}{first}:{col}{last})"
        for col in ("C", "G"):
            ws[f"{col}{r}"] = var.cell(f"=SUM({col}{first}:{col}{last})")
        ws[f"D{r}"] = var.cell(f"=B{r}-C{r}"); ws[f"E{r}"] = var.pct(f"D{r}", f"C{r}")
        ws[f"H{r}"] = var.cell(f"=F{r}-G{r}"); ws[f"I{r}"] = var.pct(f"H{r}", f"G{r}")
        ws[f"J{r}"] = var.fu(f"D{r}", not spend_only)
        num, pct_col, fu_col, ncols = ("B", "C", "D", "F", "G", "H"), "E", "J", 11
        ws[f"I{r}"].number_format = PCT; ws[f"I{r}"].font = F_SUB
        ws[f"I{r}"].alignment = RIGHT
    else:
        ws[f"B{r}"] = f"=SUM(B{first}:B{last})"
        ws[f"C{r}"] = var.cell(f"=SUM(C{first}:C{last})")
        ws[f"D{r}"] = var.cell(f"=B{r}-C{r}"); ws[f"E{r}"] = var.pct(f"D{r}", f"C{r}")
        ws[f"F{r}"] = var.fu(f"D{r}", not spend_only)
        num, pct_col, fu_col, ncols = ("B", "C", "D"), "E", "F", 7
    for col in num:
        ws[f"{col}{r}"].number_format = CUR; ws[f"{col}{r}"].font = F_SUB
        ws[f"{col}{r}"].alignment = RIGHT
    ws[f"{pct_col}{r}"].number_format = PCT; ws[f"{pct_col}{r}"].font = F_SUB
    ws[f"{pct_col}{r}"].alignment = RIGHT
    ws[f"{fu_col}{r}"].alignment = CENTER; ws[f"{fu_col}{r}"].font = F_SUB
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c); cell.fill = FILL_IVORY; cell.border = SUBTOTAL_TOP
    ws.row_dimensions[r].height = 24

    flag_rng = None if revenue_split else (f"K{first}:K{last}" if ytd else f"G{first}:G{last}")
    badge_cf(ws, f"{fu_col}{first}:{fu_col}{last}", flag_rng)
    if var:
        variance_cf(ws, f"D{first}:E{last}", f"${fu_col}{first}")
    widths(ws, [26, 15, 15, 15, 15, 14, 9, 6] if revenue_split
                else [30, 13, 13, 12, 9, 13, 13, 12, 9, 6, 11] if ytd
                else [30, 15, 15, 15, 10, 6, 11, 2])
    quiet_indicators(ws, 5, last + 2)
    ws.freeze_panes = f"A{first}"


def _group_expense_types(values: list[str]) -> list[tuple[str, list[str]]]:
    """Arrange the expense types present into readable groups.

    Known types follow the standard order (personnel first, then other operating
    costs, cost of sales, non-cash and financing). Anything a client calls by its
    own name still gets reported, collected under "Other expense types" so no
    line is silently dropped.
    """
    remaining = list(values)
    groups: list[tuple[str, list[str]]] = []
    for name, types in EXPENSE_GROUPS:
        present = [t for t in types if t in remaining]
        if present:
            groups.append((name, present))
            remaining = [t for t in remaining if t not in present]
    if remaining:
        groups.append(("Other expense types", sorted(remaining)))
    return groups


def _expense_sheet(ws, groups, gl, glf, gll, L, period, var, perno=None, ytd=False):
    """Expense report by type, grouped with subtotals, month and optional YTD."""
    hide_grid(ws)
    meta = (f"Reporting month {period}  ·  with year to date  ·  \u20ac" if ytd
            else f"Reporting month {period}  ·  \u20ac")
    title_band(ws, "Expense Report · by expense type", meta, "K" if ytd else "G")
    R = lambda c: f"'{gl}'!${c}${glf}:${c}${gll}"
    dim, cat = R(L["expense_type"]), R(L["category"])
    act, bud = R(L["actual"]), R(L["budget"])
    mflt = f',{R(L["period_no"])},{perno}' if perno and "period_no" in L else ""
    yflt = f',{R(L["period_no"])},"<={perno}"' if perno and "period_no" in L else ""

    heads = (["Expense type", "Month Act", "Month Bud", "Month Var", "Var %",
              "YTD Act", "YTD Bud", "YTD Var", "Var %", "F/U", "Flag"] if ytd else
             ["Expense type", "Actual", "Budget", "Var (Bud)", "Var %", "F/U", "Flag"])
    headers(ws, 5, heads, center_from=2, center_to=len(heads))
    fu_col, flag_col = ("J", "K") if ytd else ("F", "G")
    ncols = 11 if ytd else 7

    def write_row(r, label, formulas, bold, fill, border=False, level=0):
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = F_SUB if bold else F_BODY
        cell.alignment = indent(level) if level else LEFT
        for ref, val in formulas.items():
            ws[f"{ref}{r}"] = val
        money = ("B", "C", "D", "F", "G", "H") if ytd else ("B", "C", "D")
        pcts = ("E", "I") if ytd else ("E",)
        for col in money:
            c = ws[f"{col}{r}"]; c.number_format = CUR
            c.font = F_SUB if bold else F_BODY; c.alignment = RIGHT
        for col in pcts:
            c = ws[f"{col}{r}"]; c.number_format = PCT
            c.font = F_SUB if bold else F_BODY; c.alignment = RIGHT
        ws[f"{fu_col}{r}"].alignment = CENTER
        ws[f"{fu_col}{r}"].font = F_FU
        ws[f"{flag_col}{r}"].alignment = CENTER
        ws[f"{flag_col}{r}"].font = F_FLAG
        for c in range(1, ncols + 1):
            cc = ws.cell(row=r, column=c); cc.fill = fill
            if border:
                cc.border = SUBTOTAL_TOP
        ws.row_dimensions[r].height = 21 if bold else 18

    def line_formulas(r, crit):
        f = {"B": f'=SUMIFS({act},{cat},"<>Revenue"{crit}{mflt})',
             "C": var.cell(f'=SUMIFS({bud},{cat},"<>Revenue"{crit}{mflt})'),
             "D": var.cell(f"=B{r}-C{r}"), "E": var.pct(f"D{r}", f"C{r}")}
        if ytd:
            f |= {"F": f'=SUMIFS({act},{cat},"<>Revenue"{crit}{yflt})',
                  "G": var.cell(f'=SUMIFS({bud},{cat},"<>Revenue"{crit}{yflt})'),
                  "H": var.cell(f"=F{r}-G{r}"), "I": var.pct(f"H{r}", f"G{r}")}
        f[fu_col] = var.fu(f"D{r}", False)
        f[flag_col] = var.flag(f"D{r}", f"E{r}", PL_E, PL_P)
        return f

    def sum_formulas(r, rows):
        actual_cols = ("B", "F") if ytd else ("B",)
        budget_cols = ("C", "G") if ytd else ("C",)
        f = {c: "=" + "+".join(f"{c}{x}" for x in rows) for c in actual_cols}
        f |= {c: var.cell("=" + "+".join(f"{c}{x}" for x in rows)) for c in budget_cols}
        f |= {"D": var.cell(f"=B{r}-C{r}"), "E": var.pct(f"D{r}", f"C{r}")}
        if ytd:
            f |= {"H": var.cell(f"=F{r}-G{r}"), "I": var.pct(f"H{r}", f"G{r}")}
        f[fu_col] = var.fu(f"D{r}", False)
        f[flag_col] = var.flag(f"D{r}", f"E{r}", PL_E, PL_P)
        return f

    first, r, group_rows = 6, 6, []
    for gname, types in groups:
        # A group holding one type needs no separate subtotal line.
        single = len(types) == 1
        type_rows = []
        for i, et in enumerate(types):
            fill = FILL_IVORY if single else (band_fill(i))
            write_row(r, et, line_formulas(r, f',{dim},"{et}"'),
                      bold=single, fill=fill, border=single, level=0 if single else 1)
            type_rows.append(r); r += 1
        if single:
            group_rows.append(type_rows[0])
            continue
        write_row(r, gname, sum_formulas(r, type_rows), bold=True,
                  fill=FILL_IVORY, border=True)
        group_rows.append(r); r += 1
    last = r - 1

    write_row(r, "Total expenses", sum_formulas(r, group_rows), bold=True,
              fill=FILL_IVORY, border=True)
    for c in range(1, ncols + 1):
        ws.cell(row=r, column=c).border = TOTAL_TOP

    badge_cf(ws, f"{fu_col}{first}:{fu_col}{r}", f"{flag_col}{first}:{flag_col}{r}")
    if var:
        spend_variance_cf(ws, f"D{first}:E{r}")
        if ytd:
            spend_variance_cf(ws, f"H{first}:I{r}")
    ws.cell(row=r + 2, column=1,
            value="Grouped as a cost owner reads them: personnel first, then other "
                  "operating costs, cost of sales, and non-cash and financing items. "
                  "Bold rows are group subtotals.").font = F_NOTE
    widths(ws, [30, 13, 13, 12, 9, 13, 13, 12, 9, 6, 11] if ytd
                else [30, 15, 15, 15, 10, 6, 11])
    quiet_indicators(ws, 5, last + 3)
    ws.freeze_panes = f"B{first}"


# ---------------------------------------------------------------------------
# Departments & cost centres (hierarchy read from the data)
# ---------------------------------------------------------------------------
def _departments(ws, hierarchy, gl, glf, gll, L, period, has_cc, var, perno=None):
    hide_grid(ws)
    title_band(ws, "Departments & Cost Centres · spend variance",
                   f"Reporting month {period}  ·  \u20ac", "G")
    headers(ws, 5, ["Department / Cost centre", "Actual", "Budget", "Var (Bud)",
                        "Var %", "F/U", "Flag"], center_from=2, center_to=7)
    R = lambda c: f"'{gl}'!${c}${glf}:${c}${gll}"
    dep, cat = R(L["department"]), R(L["category"])
    act, bud = R(L["actual"]), R(L["budget"])
    cc = R(L["cost_centre"]) if has_cc else None
    mflt = f',{R(L["period_no"])},{perno}' if perno and "period_no" in L else ""

    first, r, dept_rows = 6, 6, []
    for dept, ccs in hierarchy.items():
        B, C, D, E, Fc, G = f"B{r}", f"C{r}", f"D{r}", f"E{r}", f"F{r}", f"G{r}"
        ws.cell(row=r, column=1, value=dept).font = F_SUB
        ws.cell(row=r, column=1).alignment = LEFT
        ws[B] = f'=SUMIFS({act},{dep},"{dept}",{cat},"<>Revenue"{mflt})'
        ws[C] = var.cell(f'=SUMIFS({bud},{dep},"{dept}",{cat},"<>Revenue"{mflt})')
        ws[D] = var.cell(f"={B}-{C}"); ws[E] = var.pct(D, C)
        ws[Fc] = var.fu(D, False); ws[G] = var.flag(D, E, PL_E, PL_P)
        for col in (B, C, D):
            ws[col].number_format = CUR; ws[col].font = F_SUB; ws[col].alignment = RIGHT
        ws[E].number_format = PCT; ws[E].font = F_SUB; ws[E].alignment = RIGHT
        ws[Fc].alignment = CENTER; ws[Fc].font = F_FU
        ws[G].alignment = CENTER; ws[G].font = F_FLAG
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = FILL_IVORY
            ws.cell(row=r, column=c).border = SUBTOTAL_TOP
        ws.row_dimensions[r].height = 22
        dept_rows.append(r); r += 1

        if has_cc:
            for i, c_name in enumerate(ccs):
                B, C, D, E, Fc, G = f"B{r}", f"C{r}", f"D{r}", f"E{r}", f"F{r}", f"G{r}"
                lab = ws.cell(row=r, column=1, value=str(c_name))
                lab.font = F_BODY
                lab.alignment = indent(2)
                ws[B] = f'=SUMIFS({act},{cc},"{c_name}",{cat},"<>Revenue"{mflt})'
                ws[C] = var.cell(f'=SUMIFS({bud},{cc},"{c_name}",{cat},"<>Revenue"{mflt})')
                ws[D] = var.cell(f"={B}-{C}"); ws[E] = var.pct(D, C)
                ws[Fc] = var.fu(D, False); ws[G] = var.flag(D, E, PL_E, PL_P)
                band = band_fill(i)
                for col in (B, C, D):
                    ws[col].number_format = CUR; ws[col].font = F_BODY; ws[col].alignment = RIGHT
                ws[E].number_format = PCT; ws[E].font = F_BODY; ws[E].alignment = RIGHT
                ws[Fc].alignment = CENTER
                ws[Fc].font = F_FU
                ws[G].alignment = CENTER
                ws[G].font = F_FLAG
                for c in range(1, 8):
                    ws.cell(row=r, column=c).fill = band
                ws.row_dimensions[r].height = 18
                r += 1
    last = r - 1

    ws.cell(row=r, column=1, value="Total spend").font = F_SUB
    ws[f"B{r}"] = "=" + "+".join(f"B{d}" for d in dept_rows)
    ws[f"C{r}"] = var.cell("=" + "+".join(f"C{d}" for d in dept_rows))
    ws[f"D{r}"] = var.cell(f"=B{r}-C{r}"); ws[f"E{r}"] = var.pct(f"D{r}", f"C{r}")
    ws[f"F{r}"] = var.fu(f"D{r}", False)
    for col in ("B", "C", "D"):
        ws[f"{col}{r}"].number_format = CUR; ws[f"{col}{r}"].font = F_SUB
        ws[f"{col}{r}"].alignment = RIGHT
    ws[f"E{r}"].number_format = PCT; ws[f"E{r}"].font = F_SUB; ws[f"E{r}"].alignment = RIGHT
    ws[f"F{r}"].alignment = CENTER; ws[f"F{r}"].font = F_SUB
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c); cell.fill = FILL_IVORY
        cell.border = TOTAL_TOP
    ws.row_dimensions[r].height = 24
    badge_cf(ws, f"F{first}:F{last}", f"G{first}:G{last}")
    if var:
        variance_cf(ws, f"D{first}:E{r}", f"$F{first}")
    ws.cell(row=r + 2, column=1,
            value="Departments are roll-ups; cost centres are indented beneath them. "
                  "Revenue is excluded: this is a spend view.").font = F_NOTE
    widths(ws, [32, 15, 15, 15, 10, 6, 11])
    fit_text_columns(ws, ["A"], first, r)
    quiet_indicators(ws, 5, last + 3)
    ws.freeze_panes = f"A{first}"


# ---------------------------------------------------------------------------
# Drivers
# ---------------------------------------------------------------------------
def _drivers(ws, agg, gl, glf, gll, L, period, var, perno=None):
    hide_grid(ws)
    title_band(ws, "Variance Drivers · account level", f"Reporting month {period}  ·  \u20ac", "I")
    headers(ws, 5, ["Account", "Account name", "Category", "Actual", "Budget",
                        "Var (Bud)", "Var %", "F/U", "Flag"], center_from=4, center_to=9)
    R = lambda c: f"'{gl}'!${c}${glf}:${c}${gll}"
    code_r, act, bud = R(L["account_code"]), R(L["actual"]), R(L["budget"])
    mflt = f',{R(L["period_no"])},{perno}' if perno and "period_no" in L else ""
    keep = agg
    if perno and "period_no" in agg.columns:
        keep = agg[agg["period_no"] == perno]
    tmp = (keep.groupby(["account_code", "account_name", "category"], as_index=False)
               [["actual", "budget"]].sum())
    tmp["_var"] = tmp["actual"] - tmp["budget"]
    tmp = tmp.sort_values("_var", key=lambda s: s.abs(), ascending=False)

    first, r = 6, 6
    for i, (_, row) in enumerate(tmp.iterrows()):
        favdir = CATEGORY_FAVOURABLE.get(row["category"], FAV_HIGHER)
        code = str(row["account_code"])
        band = band_fill(i)
        D, E, Fc, G, H, I = f"D{r}", f"E{r}", f"F{r}", f"G{r}", f"H{r}", f"I{r}"
        ws.cell(row=r, column=1, value=str(row["account_code"])).font = F_BODY
        ws.cell(row=r, column=2, value=str(row["account_name"])).font = F_BODY
        ws.cell(row=r, column=3, value=str(row["category"])).font = F_BODY
        ws[D] = f'=SUMIFS({act},{code_r},"{code}"{mflt})'
        ws[E] = var.cell(f'=SUMIFS({bud},{code_r},"{code}"{mflt})')
        for col in (D, E):
            ws[col].number_format = CUR; ws[col].font = F_BODY; ws[col].alignment = RIGHT
        ws[Fc] = var.cell(f"={D}-{E}"); ws[Fc].number_format = CUR; ws[Fc].font = F_BODY
        ws[Fc].alignment = RIGHT
        ws[G] = var.pct(Fc, E); ws[G].number_format = PCT; ws[G].font = F_BODY
        ws[G].alignment = RIGHT
        ws[H] = var.fu(Fc, favdir == FAV_HIGHER)
        ws[H].alignment = CENTER; ws[H].font = F_FU
        ws[I] = var.flag(Fc, G, PL_E, PL_P)
        ws[I].alignment = CENTER; ws[I].font = F_FLAG
        for c in range(1, 10):
            ws.cell(row=r, column=c).fill = band
            if c <= 3:
                ws.cell(row=r, column=c).alignment = LEFT
        ws.row_dimensions[r].height = 18
        r += 1
    last = r - 1
    ws.conditional_formatting.add(f"F{first}:F{last}",
        DataBarRule(start_type="min", end_type="max", color=BAR, showValue=True))
    badge_cf(ws, f"H{first}:H{last}", f"I{first}:I{last}")
    if var:
        variance_cf(ws, f"F{first}:G{last}", f"$H{first}")
    widths(ws, [12, 30, 12, 15, 15, 15, 10, 6, 11])
    fit_text_columns(ws, ["B", "C"], first, last)
    quiet_indicators(ws, 5, last)
    ws.freeze_panes = f"A{first}"


# ---------------------------------------------------------------------------
def build_client_pack(agg: pd.DataFrame, period: str | None = None,
                      out_path: str | Path = "flux_pack.xlsx",
                      budgeted: bool | None = None) -> Path:
    """Build the pack, adapting to the dimensions and periods in the data.

    The reporting month is the latest period present; when the file spans more
    than one period the pack also shows year-to-date beside it, which is how a
    management report is read.

    With no budget in the data the variance, F/U and materiality columns are
    left empty rather than measured against zero, and the pack says so.
    """
    from .. import ingest

    agg = agg.copy()
    for col in ("actual", "budget", "prior_year"):
        if col not in agg.columns:
            agg[col] = 0.0
        agg[col] = agg[col].fillna(0.0)
    for col in ("expense_type", "department", "cost_centre", "entity"):
        if col in agg.columns:
            agg[col] = agg[col].fillna("").astype(str).str.strip()

    # Reporting period follows from the data.
    has_period = "period" in agg.columns and agg["period"].astype(str).str.strip().ne("").any()
    all_periods = []
    perno = None
    if has_period:
        if "period_no" not in agg.columns:
            agg["period_no"] = agg["period"].map(ingest.period_key)
        latest, all_periods = ingest.reporting_period(agg)
        period = period or latest
        perno = ingest.period_key(period)
    period = period or "current period"
    ytd = has_period and len(all_periods) > 1 and perno is not None

    dims = [k for k, _l, _w in OPTIONAL_DIMS if _present(agg, k)]
    gl_dims = (["period"] if has_period else []) + dims

    # The ledger reads chronologically, then by account.
    sort_cols = (["period_no"] if "period_no" in agg.columns else []) + ["account_code"]
    agg = agg.sort_values(sort_cols, kind="stable").reset_index(drop=True)

    # Figures behind the commentary and KPI colouring use the reporting month.
    month = agg
    if perno is not None and "period_no" in agg.columns:
        month = agg[agg["period_no"] == perno]
        if month.empty:
            month = agg
    budgeted = _detect_budget(agg) if budgeted is None else bool(budgeted)
    var = Variance(budgeted)

    report = build_report(month, budgeted=budgeted)
    comments = line_comments(report, month)

    wb = Workbook()
    ws_gl = wb.active; ws_gl.title = "GL Input"
    glf, gll, L = _gl_input(ws_gl, agg, period, gl_dims, budgeted)
    _pnl(wb.create_sheet("P&L Report"), "GL Input", glf, gll, L, period,
         comments, report, var, perno, ytd)

    order = ["P&L Report"]
    if "expense_type" in dims:
        vals = [v for v in month.loc[month["category"] != "Revenue", "expense_type"].unique()
                if v and v != "(multiple)"]
        if vals:
            _expense_sheet(wb.create_sheet("Expense Report"), _group_expense_types(vals),
                           "GL Input", glf, gll, L, period, var, perno, ytd)
            order.append("Expense Report")
    if "entity" in dims:
        vals = [v for v in month["entity"].unique() if v and v != "(multiple)"]
        if len(vals) > 1:
            _dimension_sheet(wb.create_sheet("By Entity"), sorted(vals), "GL Input",
                             glf, gll, L, period, "By Entity · consolidation",
                             "Entity", "entity", var, revenue_split=True, perno=perno)
            order.append("By Entity")
    if "department" in dims:
        spend = month[month["category"] != "Revenue"]
        has_cc = "cost_centre" in dims
        hierarchy = {}
        for dept in sorted(v for v in spend["department"].unique() if v and v != "(multiple)"):
            ccs = (sorted(v for v in spend.loc[spend["department"] == dept, "cost_centre"].unique()
                          if v and v != "(multiple)") if has_cc else [])
            hierarchy[dept] = ccs
        if hierarchy:
            _departments(wb.create_sheet("Departments & CCs"), hierarchy, "GL Input",
                         glf, gll, L, period, has_cc and any(hierarchy.values()),
                         var, perno)
            order.append("Departments & CCs")

    _drivers(wb.create_sheet("Drivers"), agg, "GL Input", glf, gll, L, period, var, perno)
    order += ["Drivers", "GL Input"]
    for i, name in enumerate(order):
        wb.move_sheet(name, -wb.sheetnames.index(name) + i)
    wb.active = wb.sheetnames.index("P&L Report")

    out_path = Path(out_path); out_path.parent.mkdir(parents=True, exist_ok=True)
    quiet = collect_quiet_ranges(wb)
    wb.save(out_path)
    suppress_error_indicators(out_path, quiet)
    return out_path


if __name__ == "__main__":  # pragma: no cover - manual run
    from .. import ingest
    root = Path(__file__).resolve().parents[3]
    std, _report, _issues = ingest.ingest(root / "data" / "input_template.xlsx")
    out = build_client_pack(ingest.aggregate_to_accounts(std),
                            out_path=root / "output" / "flux_client_pack.xlsx")
    print(f"Written: {out}")
