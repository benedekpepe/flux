"""
Build the fillable Excel input template (data/input_template.xlsx).

Two sheets:
  Input        - styled header, three clearly marked example rows to delete,
                 a dropdown on Category, blue input cells.
  How to fill  - which columns are required, formats accepted, and a note that
                 the template is optional (Flux maps arbitrary headers anyway).
"""

from __future__ import annotations
from pathlib import Path

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Segoe UI"
NAVY = "1B2438"; BRASS = "C6A15B"; IVORY = "F5F1E8"; INK = "23272E"; MUTE = "8A8F98"

F_TITLE = Font(name=FONT, size=18, bold=True, color="FFFFFF")
F_SUB = Font(name=FONT, size=10, color="D7DCE6")
F_HEAD = Font(name=FONT, size=9, bold=True, color="FFFFFF")
F_BODY = Font(name=FONT, size=10, color=INK)
F_INPUT = Font(name=FONT, size=10, color="1F5FBF")
F_EX = Font(name=FONT, size=10, italic=True, color="9AA0A6")
F_NOTE = Font(name=FONT, size=9, color=MUTE)
F_H2 = Font(name=FONT, size=11, bold=True, color=NAVY)

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_BRASS = PatternFill("solid", fgColor=BRASS)
FILL_EX = PatternFill("solid", fgColor="FFF7DE")
FILL_IVORY = PatternFill("solid", fgColor=IVORY)

CUR = '#,##0.00;(#,##0.00);-'
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")

def prose_height(text: str, chars_per_line: int, *, floor: int = 16) -> float:
    """The height a wrapped, merged block of text actually needs.

    Excel auto-fits a wrapped row only when no height is set - and these rows do
    set one, to keep the sheet's rhythm. openpyxl cannot measure text, so the
    height was written by hand for each block and every later edit to the copy
    left it clipped. Three of them were, including the one a reader meets first.
    """
    lines = max(1, -(-len(str(text)) // max(20, chars_per_line)))
    return max(floor, 10 + 15 * lines)


#: Characters that fit on one line of the merged A:C block (20 + 18 + 74).
HOWTO_WIDTH = 105
#: And of the notes column on its own.
NOTES_WIDTH = 72

COLUMNS = [
    ("period", "Accounting period", False, "The period the line belongs to, e.g. 2025-06. Fill it if the file covers more than one month; Flux then reports one period at a time."),
    ("account_code", "Account code", True, "Text or number, e.g. 4100."),
    ("account_name", "Account name", True, "Free text description."),
    ("category", "Category", False, "Revenue / COGS / OpEx / Other. Left blank, Flux works it out from the account code and the account name together, and reports what it assumed."),
    ("expense_type", "Expense type", False, "Natural classification - pick from the dropdown, or use your own names. Drives the Expense Report sheet. Left blank, Flux infers it from the account name, which is the weakest guess it makes."),
    ("entity", "Entity", False, "Legal entity or company code. With more than one, the pack adds a consolidation sheet."),
    ("department", "Department", False, "The roll-up level, e.g. Engineering."),
    ("cost_centre", "Cost centre", False, "The cost centre inside that department, in your own coding scheme, e.g. CC1200, EN40300 or 4100-02. Must belong to the department in the same row."),
    ("actual", "Actual", "amount", "Actual amount for the reporting period. Fill either this or Budget - a budget-only file is a valid upload, and the app takes one as a second file."),
    ("budget", "Budget", "amount", "Budgeted amount for the same period. Fill either this or Actual. Leave the column out entirely if the plan lives in a separate file."),
    ("prior_year", "Prior year actual", False, "The same period last year, actual (not last year's budget)."),
]

EXAMPLES = [
    # Revenue posted to a commercial cost centre, as most systems do.
    ("2025-06", "4100", "Subscription revenue", "Revenue", "Revenue",
     "Flux GmbH", "Commercial", "CM10100", 512000, 480000, 445000),
    # Cost rows: the cost centre must sit inside the department on the same row.
    ("2025-06", "6110", "Advertising & digital", "OpEx", "Marketing & advertising",
     "Flux GmbH", "Marketing", "MK30200", 128000, 90000, 82000),
    ("2025-06", "6200", "Salaries - Engineering", "OpEx", "Salaries & wages",
     "Flux Inc.", "Engineering", "EN40300", 104000, 100000, 96000),
]


def build(path: Path) -> Path:
    wb = Workbook()

    # ---------------- Input sheet ----------------
    ws = wb.active
    ws.title = "Input"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = NAVY
    last = get_column_letter(len(COLUMNS))

    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = "FLUX · input template"
    ws["A1"].font = F_TITLE
    ws["A1"].alignment = LEFT
    for c in range(1, len(COLUMNS) + 1):
        ws.cell(row=1, column=c).fill = FILL_NAVY
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = ("One row per account for one reporting period. Delete the three example "
                "rows below.  * required  \u2020 fill either Actual or Budget")
    ws["A2"].font = Font(name=FONT, size=9, color=MUTE)
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 16

    ws.merge_cells(f"A3:{last}3")
    for c in range(1, len(COLUMNS) + 1):
        ws.cell(row=3, column=c).fill = FILL_BRASS
    ws.row_dimensions[3].height = 3

    hrow = 5
    for i, (key, label, required, _) in enumerate(COLUMNS, start=1):
        # "*" is required outright; "†" is one-of-these - marking Actual as
        # required outright was wrong, because the engine only needs an amount
        # and a budget-only file is a legitimate upload.
        mark = {True: " *", "amount": " \u2020"}.get(required, "")
        cell = ws.cell(row=hrow, column=i, value=label + mark)
        cell.font = F_HEAD
        cell.fill = FILL_NAVY
        cell.alignment = CENTER if i >= 9 else LEFT
        cell.border = Border(bottom=Side(style="medium", color=NAVY))
    ws.row_dimensions[hrow].height = 22

    # Example rows (amber, italic, clearly disposable).
    r = hrow + 1
    for ex in EXAMPLES:
        for c, val in enumerate(ex, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = FILL_EX
            cell.font = F_EX
            cell.alignment = RIGHT if c >= 9 else LEFT
            if c >= 9:
                cell.number_format = CUR
        ws.cell(row=r, column=len(COLUMNS) + 2, value="<- example row, delete").font = F_NOTE
        r += 1

    # Empty input rows, pre-formatted.
    for _ in range(60):
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = F_INPUT
            cell.alignment = RIGHT if c >= 9 else LEFT
            if c >= 9:
                cell.number_format = CUR
        r += 1
    last_row = r - 1

    # Category dropdown.
    dv = DataValidation(type="list", formula1='"Revenue,COGS,OpEx,Other"', allow_blank=True)
    dv.error = "Choose Revenue, COGS, OpEx or Other (or leave blank)."
    dv.prompt = "Optional - leave blank and Flux works it out from the code and the name."
    ws.add_data_validation(dv)
    dv.add(f"D{hrow+1}:D{last_row}")

    # And one for the expense types. Seventeen names typed by hand is seventeen
    # chances to split one group across two spellings.
    try:
        import sys as _sys
        _sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
        from flux.coa import EXPENSE_TYPES
        # Excel caps an inline list at 250 characters and seventeen type names
        # run well past that, so the list lives on a hidden sheet and the
        # validation points at it.
        lists = wb.create_sheet("Lists")
        lists["A1"] = "Expense types - the list behind the dropdown on Input."
        for i, etype in enumerate(EXPENSE_TYPES, start=2):
            lists.cell(row=i, column=1, value=etype)
        lists.column_dimensions["A"].width = 30
        lists.sheet_state = "hidden"
        ref = f"Lists!$A$2:$A${len(EXPENSE_TYPES) + 1}"
        dv2 = DataValidation(type="list", formula1=ref, allow_blank=True)
        # A warning, not a hard stop: a client's own expense names are valid
        # input and get grouped under "Other expense types" rather than refused.
        dv2.errorStyle = "warning"
        dv2.error = "Not one of the standard types - your own names are fine too."
        dv2.prompt = "Pick one, or type your own name."
        ws.add_data_validation(dv2)
        dv2.add(f"E{hrow+1}:E{last_row}")
    except Exception as exc:  # pragma: no cover
        print(f"WARNING: expense-type dropdown not added ({exc!r})")

    for i, (_, _, _, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = [17, 14, 30, 13, 24, 15, 16, 15, 15, 15, 16][i - 1]
    ws.column_dimensions[get_column_letter(len(COLUMNS) + 2)].width = 26
    ws.freeze_panes = f"A{hrow+1}"

    # ---------------- How to fill sheet ----------------
    ws2 = wb.create_sheet("How to fill")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = NAVY
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "How to fill this template"
    ws2["A1"].font = F_TITLE
    ws2["A1"].alignment = LEFT
    for c in range(1, 4):
        ws2.cell(row=1, column=c).fill = FILL_NAVY
    ws2.row_dimensions[1].height = 34
    ws2.merge_cells("A2:C2")
    for c in range(1, 4):
        ws2.cell(row=2, column=c).fill = FILL_BRASS
    ws2.row_dimensions[2].height = 3

    ws2["A4"] = "Why this template is the recommended way in"
    ws2["A4"].font = F_H2
    ws2.merge_cells("A5:C5")
    ws2["A5"] = ("Flux also reads general-ledger and trial-balance exports as they are - it "
                 "recognises column names in English, Hungarian and German, classifies the "
                 "accounts from their codes and names, and drops the subtotal rows a printed "
                 "report carries. The numbers come out the same either way when it reads the "
                 "file correctly. What this template removes is the reading: nothing to map, "
                 "no chart of accounts to recognise, no expense types inferred from account "
                 "names, and no printed totals to filter out. Use it when the pack has to be "
                 "right the first time; upload your raw export when you would rather check "
                 "four assumptions than fill a sheet.")
    ws2["A5"].alignment = WRAP
    ws2["A5"].font = F_BODY
    ws2.row_dimensions[5].height = prose_height(ws2["A5"].value, HOWTO_WIDTH)

    ws2["A7"] = "Columns"
    ws2["A7"].font = F_H2
    hdr = ["Column", "Required", "Notes"]
    for c, h in enumerate(hdr, start=1):
        cell = ws2.cell(row=8, column=c, value=h)
        cell.font = F_HEAD
        cell.fill = FILL_NAVY
        cell.alignment = LEFT
    ws2.row_dimensions[8].height = 20

    rr = 9
    for key, label, required, note in COLUMNS:
        ws2.cell(row=rr, column=1, value=label).font = F_BODY
        wording = {True: "Required", "amount": "Actual or Budget"}.get(required, "Optional")
        c2 = ws2.cell(row=rr, column=2, value=wording)
        c2.font = Font(name=FONT, size=10, bold=bool(required),
                       color=NAVY if required else MUTE)
        c3 = ws2.cell(row=rr, column=3, value=note)
        c3.font = F_BODY
        c3.alignment = WRAP
        if rr % 2:
            for c in range(1, 4):
                ws2.cell(row=rr, column=c).fill = FILL_IVORY
        # The notes column wraps, so the row has to be as tall as its own note.
        # A flat 30 clipped the longest of them mid-sentence.
        ws2.row_dimensions[rr].height = prose_height(note, NOTES_WIDTH, floor=30)
        rr += 1

    ws2.cell(row=rr + 1, column=1, value="Department and cost centre").font = F_H2
    ws2.merge_cells(f"A{rr+2}:C{rr+2}")
    h = ws2.cell(row=rr + 2, column=1)
    h.value = ("Cost centres sit inside departments; the department is the roll-up level. "
               "Put the department in one column and the cost centre it belongs to in the "
               "next, on the same row - for example Engineering / QA & Release, or "
               "Marketing / MK30200. Enter the cost centre in your own coding scheme - "
               "CC1200, EN40300 and 4100-02 are all fine, Flux takes it as it is. "
               "Post revenue the way your own system does: many companies book revenue to a "
               "commercial or sales cost centre (and credit notes always follow the cost centre "
               "of the original entry), while others track revenue by profit centre only. Flux "
               "accepts either - fill the columns if you have them, leave them blank if you do "
               "not. Every optional dimension you fill adds a sheet to the pack: expense "
               "type gives the Expense Report, department and cost centre give the "
               "Departments & Cost Centres view, and more than one entity gives a "
               "consolidation sheet. Without them you still get the P&L and the "
               "account-level variance. The departmental view reports spend, so revenue "
               "rows are excluded from it either way.")
    h.alignment = WRAP
    h.font = F_BODY
    ws2.row_dimensions[rr + 2].height = prose_height(
        ws2.cell(row=rr + 2, column=1).value, HOWTO_WIDTH)

    ws2.cell(row=rr + 4, column=1, value="Numbers and periods").font = F_H2
    ws2.merge_cells(f"A{rr+5}:C{rr+5}")
    n = ws2.cell(row=rr + 5, column=1)
    n.value = ("Amounts can use either European (1.234.567,89) or US (1,234,567.89) formatting; "
               "Flux parses both. Enter positive magnitudes: revenue as a positive number and "
               "costs as positive numbers too. Several periods in one file are fine and are "
               "what produce the year-to-date columns - fill the Accounting period column and "
               "Flux reports the latest month that carries postings. With no period column at "
               "all, the app asks what to label the report.")
    n.alignment = WRAP
    n.font = F_BODY
    ws2.row_dimensions[rr + 5].height = prose_height(n.value, HOWTO_WIDTH)

    # "Actual or Budget" needs the room: at 12 it read "Actual or B".
    for col, w in (("A", 20), ("B", 18), ("C", 74)):
        ws2.column_dimensions[col].width = w

    # ---------------- Example: full GL export ----------------
    ws3 = wb.create_sheet("Example - full GL export")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = NAVY
    # The package lives under src/, not beside this script. Importing from
    # `scripts/` failed, the bare except swallowed it, and the sheet shipped
    # with a sentence promising rows that were never written. A missing example
    # is not worth a crash, but it is worth saying out loud.
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
    try:
        from flux.synthetic_data import generate_transactions
        txns = generate_transactions("2025-06")
        # A spread across the categories, not the first twelve rows: taken in
        # file order the sample was all cost lines, so the sheet illustrated a
        # ledger with no revenue in it.
        picks = [g.head(3) for _cat, g in txns.groupby("category", sort=False)]
        sample = (pd.concat(picks).sort_values("doc_no").reset_index(drop=True)
                  if picks else txns.head(12))
    except Exception as exc:  # pragma: no cover - the sheet degrades, loudly
        print(f"WARNING: example ledger rows not generated ({exc!r})")
        sample = None

    # The bands have to span the data, not a guessed eight columns: the example
    # is a real export and runs to three dozen fields, so a title merged to H
    # was cut off mid-word and the intro clipped at its own row height.
    span = len(sample.columns) if sample is not None else 8
    last = get_column_letter(span)
    ws3.merge_cells(f"A1:{last}1")
    ws3["A1"] = "Example - what a full ledger export looks like"
    ws3["A1"].font = F_TITLE
    ws3["A1"].alignment = LEFT
    for c in range(1, span + 1):
        ws3.cell(row=1, column=c).fill = FILL_NAVY
    ws3.row_dimensions[1].height = 34
    ws3.merge_cells(f"A2:{last}2")
    for c in range(1, span + 1):
        ws3.cell(row=2, column=c).fill = FILL_BRASS
    ws3.row_dimensions[2].height = 3
    ws3.merge_cells(f"A4:{last}4")
    ws3["A4"] = ("You do not need to trim your export down to the Input sheet. Upload the "
                 "extract exactly as your system produces it - document numbers, texts, "
                 "dates, tax codes and any other fields are simply ignored. Flux only needs "
                 "the account, the amounts and (optionally) department, cost centre and "
                 "category. The rows below are an illustrative extract.")
    ws3["A4"].alignment = WRAP
    ws3["A4"].font = F_BODY
    # Excel does not auto-fit a merged cell, so the height is estimated from the
    # text and the width it has to wrap inside. At eight columns the note ran to
    # five lines in a box built for four.
    ws3.row_dimensions[4].height = prose_height(
        ws3["A4"].value, max(40, span * 11), floor=34)

    if sample is not None:
        hrow = 6
        for i, col in enumerate(sample.columns, start=1):
            cell = ws3.cell(row=hrow, column=i, value=col)
            cell.font = F_HEAD; cell.fill = FILL_NAVY; cell.alignment = LEFT
            ws3.column_dimensions[get_column_letter(i)].width = max(11, min(26, len(str(col)) + 6))
        ws3.row_dimensions[hrow].height = 20
        for r_i, (_, row) in enumerate(sample.iterrows()):
            rr = hrow + 1 + r_i
            for c_i, col in enumerate(sample.columns, start=1):
                cell = ws3.cell(row=rr, column=c_i, value=row[col])
                cell.font = Font(name=FONT, size=9, color=INK)
                cell.alignment = LEFT
                if r_i % 2:
                    cell.fill = FILL_IVORY
            ws3.row_dimensions[rr].height = 14
        ws3.freeze_panes = f"A{hrow+1}"

    wb.active = 0
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


if __name__ == "__main__":
    out = build(Path(__file__).resolve().parents[1] / "data" / "input_template.xlsx")
    print(f"Written: {out}")
